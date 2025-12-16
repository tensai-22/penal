
import logging
import os
from flask import Flask
from flask_cors import CORS
import waitress
import tempfile 
from datetime import datetime, timedelta
from PyPDF2 import PdfMerger
from pathlib import Path
from flask import jsonify, request, session
import json
from datetime import date
from dateutil import parser
import shutil
from flask import send_from_directory

# ---------------------------
# CONFIGURACIÓN DE LOGGING
# ---------------------------

# Ruta del archivo de logs
log_path = r"C:\temp\combined.log"
log_folder = os.path.dirname(log_path)

# Crear la carpeta de logs si no existe
if not os.path.exists(log_folder):
    os.makedirs(log_folder)

# Configuración de logging para toda la aplicación
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    handlers=[
        logging.FileHandler(log_path, encoding="utf-8")
    ]
)

# Obtener loggers específicos
logger = logging.getLogger(__name__)  # Logger de la aplicación Flask
werkzeug_logger = logging.getLogger("werkzeug")  # Logger de Flask/Werkzeug
waitress_logger = logging.getLogger("waitress")  # Logger de Waitress

# Configurar niveles de log para Flask y Waitress
werkzeug_logger.setLevel(logging.INFO)
waitress_logger.setLevel(logging.INFO)

# Configurar logs explícitamente en Waitress
waitress_logger.addHandler(logging.FileHandler(log_path, encoding="utf-8"))
waitress_logger.propagate = True

# Asegurar que Flask use la configuración global de logging
app = Flask(__name__)
app.secret_key = "your_secret_key"
app.logger.setLevel(logging.INFO)

# Asociar el logger de Flask con el mismo manejador de logs
for handler in logging.getLogger().handlers:
    app.logger.addHandler(handler)

CORS(
    app,
    supports_credentials=True,
    origins=[
        "http://localhost:3000",
        "http://10.50.5.49:3000",
        "http://192.168.1.42:3000",
        "http://127.0.0.1:3000"
    ]
)


#blueprints

from backend.modules.ingresos.ingresos import ingresos_bp
app.register_blueprint(ingresos_bp, url_prefix="/api")


from backend.core import (
    # Auth/decorators
    login_required, role_required,
    # DB y utilidades
    get_db_connection, allowed_file, normalize_text, query_to_regexp,
    # PPU helpers
    parse_ppu, parse_query_ppu, generar_variantes_ppu,
    # Expedientes
    normalizar_expediente,
    # PDF helpers y formato
    extract_pdf_pages, format_legajo,
    # Validación y lookups
    validate_expediente_juzgado, get_fiscalia_departamento,
)


# ---------------------------
# CONFIGURACIÓN DE ARCHIVOS
# ---------------------------

UPLOAD_FOLDER = "uploads/pdfs"
ALLOWED_EXTENSIONS = {"pdf"}

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
app.config["MAX_CONTENT_LENGTH"] = 500 * 1024 * 1024  # 500 MB

# ---------------------------
# LOGS DE INICIO
# ---------------------------

logger.info("Aplicación Flask iniciada correctamente.")
waitress_logger.info("Servidor Waitress configurado.")

# ---------------------------
# INICIO DEL SERVIDOR (EJECUCIÓN EN PRODUCCIÓN)
# ---------------------------

def run_server():
    logger.info("Iniciando el servidor en Waitress...")
    waitress.serve(app, host="0.0.0.0", port=5001)



@app.route("/test-log")
def test_log():
    logger.info("Este es un log de prueba desde una ruta Flask.")
    return "Log enviado", 200








# Ruta destino
DESTINATION_BASE_PATH = r'\\agarciaf\NOTIFICACIONES RENIEC\MESA DE PARTES\PENAL\NOTIFICACIONES'





username_to_abogado = {
    "jpolom": "POLO",
    "enavarro": "NAVARRO",
    "mpalacios": "PALACIOS",
    "imartinez": "MARTINEZ",
    "mrojas": "ROJAS",
    "mfrisancho": "FRISANCHO",
    "tpomar": "POMAR",
    "dflores": "FLORES",
    "zaguilar": "AGUILAR",
    "mmau": "MAU",
    "fascurra": "ASCURRA",
    "ncuba": "CUBA"
}

users = {
    "agarcia": {"password": "agarcia", "role": "admin"},
    "Manuel": {"password": "Manuel", "role": "admin"},
    "jgranda": {"password": "jgranda", "role": "admin"},
    # Ojo: si quieres que "abogado" sea admin en vez de user, ajusta "role": "admin".
    "abogado": {"password": "abogado", "role": "admin"},

    "jpolom":  {"password": "jpolom", "role": "user"},

    "enavarro": {"password": "enavarro", "role": "user"},
    "mpalacios": {"password": "mpalacios", "role": "user"},
    "imartinez": {"password": "imartinez", "role": "user"},
    "mrojas": {"password": "mrojas", "role": "user"},
    "mfrisancho": {"password": "mfrisancho", "role": "user"},
    "tpomar": {"password": "tpomar", "role": "user"},
    "dflores": {"password": "dflores", "role": "user"},
    "zaguilar": {"password": "zaguilar", "role": "user"},
    "mmau": {"password": "mmau", "role": "user"},
    "fascurra": {"password": "fascurra", "role": "user"},
    "ncuba": {"password": "ncuba", "role": "user"},
    "archivo-penal": {"password": "archivo2025", "role": "user"},
}





## fin de la busqueda avanzada
@app.route('/api/login', methods=['POST'])
def login():
    data = request.json
    username = data.get('username')
    password = data.get('password')

    user = users.get(username)
    if user and user['password'] == password:
        session['username'] = username
        session['role'] = user['role']
        return jsonify({"message": "Inicio de sesión exitoso", "role": user['role']})
    else:
        return jsonify({"error": "Credenciales inválidas"}), 401

@app.route('/api/logout', methods=['POST'])
def logout():
    session.clear()
    return jsonify({"message": "Cierre de sesión exitoso"})

@app.route('/api/years', methods=['GET'])
@login_required
def get_years():
    """
    Devuelve la lista de años únicos extraídos de registro_ppu,
    ordenados de mayor a menor.
    """
    app.logger.debug("→ Entrando a /api/years")  # <--- LOG

    connection = get_db_connection()
    if connection is None:
        app.logger.error("get_years: No se pudo conectar a la base de datos")
        return jsonify({"error": "Error al conectar con la base de datos"}), 500

    try:
        cursor = connection.cursor(dictionary=True)
        sql = """
            SELECT DISTINCT
              CAST(
                SUBSTRING_INDEX(
                  SUBSTRING_INDEX(registro_ppu, '-', 3),
                  '-',
                  -1
                ) AS UNSIGNED
              ) AS year
            FROM datapenal
            WHERE registro_ppu IS NOT NULL AND registro_ppu != ''
            ORDER BY year DESC
        """
        app.logger.debug("get_years: Ejecutando SQL:\n%s", sql)  # <--- LOG
        cursor.execute(sql)
        rows = cursor.fetchall()
        app.logger.debug("get_years: Filas obtenidas: %d", len(rows))  # <--- LOG

        años = [str(r["year"]) for r in rows if r.get("year") is not None]
        app.logger.debug("get_years: Lista de años resultante → %s", años)  # <--- LOG
        return jsonify({"years": años})
    except Exception as e:
        app.logger.error("Error en /api/years: %s", e, exc_info=True)
        return jsonify({"error": "Error al obtener años"}), 500
    finally:
        cursor.close()
        connection.close()

@app.route('/api/buscar', methods=['GET'])
@login_required
def buscar():
    """
    Búsqueda en `datapenal` con estos parámetros:

      - query               -> texto libre (REGEXP) (buscador global)
      - page, limit         -> paginación
      - abogado             -> filtra por abogado (admin) o forzado para user
      - mostrar_archivados  -> 'true' | 'false'
      - year                -> filtrar registros de un año concreto
      - tipo                -> 'DENUNCIA' | 'LEGAJO' | 'ALL'

    Flujo:
      1) Leo todos los parámetros.
      2) Si no se envía ni `query` ni `year`, obtengo el año más alto de la tabla
         (`MAX(CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(registro_ppu,'-',3),'-',-1) AS UNSIGNED))`).
      3) Construyo el WHERE con los filtros (abogado, mostrar_archivados, query o year+tipo).
      4) Devuelvo la lista paginada *y* el campo "used_year" indicando el año que realmente usé.
    """
    # 1) Leer parámetros de la request
    query              = request.args.get('query', '').strip()
    page               = max(int(request.args.get('page', 1)), 1)
    limit              = max(int(request.args.get('limit', 20)), 1)
    offset             = (page - 1) * limit
    mostrar_archivados = request.args.get('mostrar_archivados', 'true').lower() == 'true'
    year_param         = request.args.get('year', '').strip()      # Si viene vacío, lo asignamos luego
    tipo_param         = request.args.get('tipo', 'ALL').upper()    # 'DENUNCIA' | 'LEGAJO' | 'ALL'
    abogado_param      = request.args.get('abogado', '').strip().upper()

    app.logger.debug(
        "→ Entrando a /api/buscar con params → "
        "query: '%s', page: %d, limit: %d, mostrar_archivados: %s, year: '%s', tipo: '%s', abogado: '%s'",
        query, page, limit, mostrar_archivados, year_param, tipo_param, abogado_param
    )

        # 2) Determinar filtro de abogado según rol
    user_role    = session.get('role')
    current_user = session.get('username', '')
    if user_role == 'user':
        abogado_filter = username_to_abogado.get(current_user, '').upper()
    else:
        # Si viene "A; B", tomamos sólo "B"; si no, usamos tal cual
        raw = abogado_param
        if ';' in raw:
            abogado_filter = raw.split(';')[-1].strip().upper()
        else:
            abogado_filter = raw.upper()


    app.logger.debug(
        "   • Filtro de abogado determinado → '%s' (rol: %s, usuario: %s)",
        abogado_filter, user_role, current_user
    )

    # 3) Conectar a la base de datos
    connection = get_db_connection()
    if connection is None:
        app.logger.error("buscar: No se pudo conectar a la base de datos")
        return jsonify({"error": "Error al conectar con la base de datos"}), 500

    try:
        cursor = connection.cursor(dictionary=True)

        # 4) Si no hay query y no hay year, obtenemos el año más alto de registro_ppu
        used_year = year_param  # Lo que finalmente usaremos como filtro de año
        if not query and not year_param:
            cursor.execute("""
                SELECT
                  MAX(
                    CAST(
                      SUBSTRING_INDEX(
                        SUBSTRING_INDEX(registro_ppu, '-', 3),
                        '-',
                        -1
                      ) AS UNSIGNED
                    )
                  ) AS maxyear
                FROM datapenal
                WHERE registro_ppu IS NOT NULL AND registro_ppu != ''
            """)
            result = cursor.fetchone()
            maxyear = result.get('maxyear') if result else None
            if maxyear:
                used_year = str(maxyear)
                app.logger.debug(
                    "   • No se envió 'year' ni 'query': asigno año más alto → %s",
                    used_year
                )

        # 5) Construir condiciones dinámicas
        conditions = []
        params     = []

        # 5.1 · Excluir archivados si corresponde
        if not mostrar_archivados:
            conditions.append("(d.etiqueta IS NULL OR d.etiqueta != %s)")
            params.append('ARCHIVO')

         # 5.2 · Filtrar por abogado usando sólo la parte posterior a ';'
        if abogado_filter:
            # Extrae la parte tras el ';' (o toda la cadena si no hay ';'), la limpia y compara en mayúsculas
            conditions.append(
                "UPPER(TRIM(SUBSTRING_INDEX(d.abogado, ';', -1))) = %s"
            )
            params.append(abogado_filter)

        # 5.3 · Si hay query (buscador global), ignoramos filtros de year/tipo
        if query:
            cols    = [
                'abogado','denunciado','origen','delito','departamento',
                'fiscalia','informe_juridico','item','e_situacional',
                'registro_ppu','juzgado','etiqueta'
            ]
            regexp    = query_to_regexp(query)
            sub_conds = [f"d.{c} REGEXP %s" for c in cols]
            conditions.insert(0, "(" + " OR ".join(sub_conds) + ")")
            params = [regexp] * len(cols) + params
            app.logger.debug(
                "   • Buscador global: query='%s' → aplicar REGEXP en columnas: %s",
                query, cols
            )

        else:
            # 5.4 · Cuando NO hay query, aplicamos un solo filtro REGEXP según tipo y año
            # … dentro de la función buscar(), cuando no hay query:
            if not query and used_year and used_year.isdigit():
                if tipo_param == 'DENUNCIA':
                    # 1) “D-<número>-<año>” o “D-<número>-<año>-<letra>”
                    #    => [0-9]+ acepta cualquier cantidad de dígitos en el número, y (?:-[A-Z])? el sufijo opcional.
                    pattern = rf"^D-[0-9]+-{re.escape(used_year)}(?:-[A-Z])?$"
                    app.logger.debug("   • Filtro unificado: tipo='DENUNCIA', pattern = '%s'", pattern)

                elif tipo_param == 'LEGAJO':
                    # 2) “LEG-<número>-<año>” o “LEG-<número>-<año>-<letra>”
                    #    ó “L. <número>-<año>” o “L. <número>-<año>-<letra>”
                    pattern = rf"^(?:LEG-[0-9]+-{re.escape(used_year)}(?:-[A-Z])?|L\. ?[0-9]+-{re.escape(used_year)}(?:-[A-Z])?)$"
                    app.logger.debug("   • Filtro unificado: tipo='LEGAJO', pattern = '%s'", pattern)

                else:  # ALL
                    # 3) Fusionamos los dos casos anteriores en una sola expresión:
                    pattern = (
                        rf"^(?:"
                        # DENUNCIA: D-<número>-<año>[-<letra>]
                        rf"D-[0-9]+-{re.escape(used_year)}(?:-[A-Z])?"
                        rf"|"
                        # LEG: LEG-<número>-<año>[-<letra>]
                        rf"LEG-[0-9]+-{re.escape(used_year)}(?:-[A-Z])?"
                        rf"|"
                        # L.: L. <número>-<año>[-<letra>] (puede haber o no el punto y espacio)
                        rf"L\.?\s?[0-9]+-{re.escape(used_year)}(?:-[A-Z])?"
                        rf")$"
                    )
                    app.logger.debug("   • Filtro unificado: tipo='ALL', pattern = '%s'", pattern)

                conditions.append("d.registro_ppu REGEXP %s")
                params.append(pattern)



        # 6) Componer cláusula WHERE
        where_clause = f"WHERE {' AND '.join(conditions)}" if conditions else ""
        app.logger.debug("   • Cláusula WHERE generada → %s", where_clause)
        app.logger.debug("   • Parámetros para WHERE → %s", params)

        # 7) Ejecutar SELECT
        sql = f"""
            SELECT d.*
            FROM   datapenal d
            {where_clause}
        """
        app.logger.debug("   • Ejecutando SQL completo:\n%s", sql)
        cursor.execute(sql, params)
        rows = cursor.fetchall()

        # Limpiar el campo 'abogado': si contiene ';', quedarnos solo con la parte final
        for row in rows:
            abogado = row.get('abogado') or ''
            if ';' in abogado:
                row['abogado'] = abogado.split(';')[-1].strip()
        app.logger.debug("   • Número de filas obtenidas de la BD: %d", len(rows))

           # 8) Ordenar en Python por año descendente, tipo y número de legajo/denuncia
        def extraer_year(ppu: str) -> int:
            m = re.search(r'-(\d{4})(?:-|$)', ppu or '')
            return int(m.group(1)) if m else 0

        def tipo_order(ppu: str) -> int:
            if ppu.upper().startswith('D-'):
                return 0
            if re.match(r'^(L\. ?|LEG-)', ppu, re.IGNORECASE):
                return 1
            return 2

        def extraer_numero(ppu: str) -> int:
            """
            Extrae la parte numérica que viene después de 'D-', 'LEG-' o 'L. ' y antes del siguiente guion.
            Ejemplo: “L. 100-2024” → devuelve 100
                     “D-2501-2018-A” → devuelve 2501
            """
            m = re.search(r'^(?:D-|LEG-?|L\.?\s?)(\d+)-', ppu or '', re.IGNORECASE)
            return int(m.group(1)) if m else 0

        def sort_key(r):
            ppu = (r.get('registro_ppu') or '').upper()
            y   = extraer_year(ppu)
            t   = tipo_order(ppu)
            n   = extraer_numero(ppu)
            return (-y, t, n)   # <<< ahora usamos 'n' en lugar de 'ppu' para ordenar numéricamente

        rows_sorted   = sorted(rows, key=sort_key)
        total_records = len(rows_sorted)
        total_pages   = (total_records + limit - 1) // limit
        data_page     = rows_sorted[offset : offset + limit]

        app.logger.debug(
            "   • Total registros ordenados: %d, páginas: %d, offset: %d",
            total_records, total_pages, offset
        )

        # 9) Reemplazar None por '' y devolver JSON
        # 9) Limpiar el campo 'abogado' y luego reemplazar None por ''
        for fila in data_page:
            # si viene formato "X; Y", nos quedamos sólo con Y
            abogado = fila.get('abogado', '') or ''
            if ';' in abogado:
                fila['abogado'] = abogado.split(';')[-1].strip().upper()

            # luego, el resto de campos None → ""
            for k, v in fila.items():
                if v is None:
                    fila[k] = ""


        respuesta = {
            "data"          : data_page,
            "page"          : page,
            "total_pages"   : total_pages,
            "total_records" : total_records,
            "used_year"     : used_year  # <-- Este es el año que el backend realmente aplicó
        }
        app.logger.debug(
            "   • Respuesta JSON de /api/buscar → keys: %s",
            respuesta.keys()
        )
        return jsonify(respuesta)

    except Exception as e:
        app.logger.error("Error en /api/buscar: %s", e, exc_info=True)
        return jsonify({"error": "Error al realizar la búsqueda"}), 500

    finally:
        cursor.close()
        connection.close()


from io import BytesIO
from datetime import datetime
from collections import defaultdict
import difflib
import pandas as pd
from flask import request, jsonify, session, make_response


@app.route('/api/exportar_excel', methods=['GET'])
@login_required
def exportar_excel():
    """
    Exporta registros de datapenal a Excel con:
      - Limpieza de 'abogado' (solo parte tras ';')
      - Filtros (ppu rango, búsqueda global, filtro por abogado)
      - Métricas por año y totales (solo del abogado filtrado o de todos si admin sin filtro)
      - Formato avanzado: tablas de Excel, centrado, ancho fijo, filas prohibidas en rojo
    """
    # 1) Parámetros
    ppu_inicio         = request.args.get('ppu_inicio', '').strip()
    ppu_fin            = request.args.get('ppu_fin', '').strip()
    query              = request.args.get('query', '').strip()
    mostrar_archivados = request.args.get('mostrar_archivados','true').lower() == 'true'
    tipo               = request.args.get('tipo', 'ALL').strip().upper()

    # 2) Rol y filtro de abogado
    user_role    = session.get('role')
    current_user = session.get('username','')
    if user_role == 'user':
        abogado_filter = username_to_abogado.get(current_user,'').upper()
    else:
        abogado_filter = request.args.get('abogado','').strip().upper() or ''

    # 3) Conexión
    conn = get_db_connection()
    if conn is None:
        return jsonify({"error":"Error al conectar con la base de datos"}), 500

    try:
        cursor = conn.cursor(dictionary=True)

        # 4) WHERE dinámico
        conds, params = [], []
        if not mostrar_archivados:
            conds.append("(etiqueta IS NULL OR etiqueta<>%s)"); params.append('ARCHIVO')
        if abogado_filter:
            conds.append("UPPER(TRIM(SUBSTRING_INDEX(abogado,';',-1)))=%s"); params.append(abogado_filter)
        if query:
            cols = ['abogado','denunciado','origen','delito','departamento',
                    'fiscalia','informe_juridico','item','e_situacional',
                    'registro_ppu','juzgado','etiqueta']
            rx = query_to_regexp(query)
            sub = []
            for c in cols:
                sub.append(f"{c} REGEXP %s"); params.append(rx)
            conds.append("("+ " OR ".join(sub)+")")
        where = "WHERE " + " AND ".join(conds) if conds else ""

        # 5) Fetch datos
        cursor.execute(f"SELECT * FROM datapenal {where}", params)
        rows = cursor.fetchall()

        # 6) Orden inicial
        rows.sort(key=lambda r: parse_ppu(r['registro_ppu']))

        # 7) Filtrar por rango
        start_t = parse_ppu(ppu_inicio) if ppu_inicio else (0,0,0,'')
        end_t   = parse_ppu(ppu_fin)    if ppu_fin    else (999999,9999,9999,'ZZZZ')
        y_min, y_max = start_t[1], end_t[1]
        final = [
            r for r in rows
            if y_min <= parse_ppu(r['registro_ppu'])[1] <= y_max
            and start_t <= parse_ppu(r['registro_ppu']) <= end_t
        ]
        # 7-bis) Filtrar por tipo (LEGAJO / DENUNCIA)
        if tipo == 'LEGAJO':
            # Excluimos todo lo que empiece por “D-”
            final = [r for r in final
                     if not r['registro_ppu'].upper().startswith('D-')]
        elif tipo == 'DENUNCIA':
            # Nos quedamos sólo con los que empiecen por “D-”
            final = [r for r in final
                     if r['registro_ppu'].upper().startswith('D-')]
        # Si tipo == 'ALL' no tocamos nada

        # 8) Limpiar abogado y marcar prohibidas
        banned_kw = ["ACUM","ACUMULADO","SUSPENDIDO","ANULADO","DERIVADO", "DUPLICADO"]
        for r in final:
            ab = (r.get('abogado') or '').split(';')[-1].strip().upper()
            r['abogado']    = ab
            r['_prohibida'] = any(kw in ab for kw in banned_kw)

        # 9) Agrupar por año
        data_by_year = defaultdict(list)
        for r in final:
            yr = parse_ppu(r['registro_ppu'])[1]
            data_by_year[yr].append(r)
        years = sorted(data_by_year.keys(), reverse=True)

        # 10) Calcular métricas
        allowed = ["CUBA","AGUILAR","POLO","MAU","ASCURRA",
                   "MARTINEZ","FLORES","PALACIOS","POMAR",
                   "ROJAS","FRISANCHO","NAVARRO"]

        # Si hay filtro de abogado, métricas solo de ese
        if abogado_filter:
            keys = [abogado_filter]
        else:
            keys = allowed + ["OTROS"]

        counts = {
            key: {str(y):0 for y in years} | {"Total":0}
            for key in keys
        }

        for r in final:
            if r['_prohibida']:
                continue
            ab = r['abogado']
            if abogado_filter and ab != abogado_filter:
                continue
            key = ab if abogado_filter else (difflib.get_close_matches(ab, allowed, n=1, cutoff=0.8) or ["OTROS"])[0]
            yr  = str(parse_ppu(r['registro_ppu'])[1])
            counts.setdefault(key, {str(y):0 for y in years}|{"Total":0})
            counts[key][yr]      += 1
            counts[key]["Total"] += 1

        metrics = []
        for key in keys:
            c = counts.get(key, {str(y):0 for y in years}|{"Total":0})
            row = {"Abogado": key}
            for y in years:
                row[str(y)] = c[str(y)]
            row["Total"] = c["Total"]
            metrics.append(row)

        # 11) Crear Excel
        output = BytesIO()
        fecha_cod = datetime.now().strftime("%d-%m-%Y %Hh%Mm")
        display = abogado_filter or "GENERAL"
        # Nombre de archivo dinámico
        filename = f"Base de datos del año {years[0] if years else ''} - {display} a la fecha de {fecha_cod}.xlsx"

        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            wb        = writer.book
            title_fmt = wb.add_format({'align':'center','valign':'vcenter','bold':True,'font_size':14})
            hdr_fmt   = wb.add_format({'align':'center','valign':'vcenter','bold':True})
            cell_fmt  = wb.add_format({'align':'center','valign':'vcenter','text_wrap':False})
            red_fmt   = wb.add_format({'bg_color':'#FFC7CE','font_color':'#9C0006'})

            # Hoja Métricas
            dfm = pd.DataFrame(metrics)
            dfm = dfm[['Abogado'] + [str(y) for y in years] + ['Total']]
            dfm.to_excel(writer, 'Métricas', startrow=1, startcol=1, index=False, header=True)
            ws = writer.sheets['Métricas']
            ws.merge_range(0,1,0,dfm.shape[1],
                           f"Métricas de {display} al {fecha_cod}", title_fmt)
            nr, nc = len(dfm), len(dfm.columns)
            ws.add_table(1,1,1+nr,1+nc-1,{
                'columns':[{'header':h} for h in dfm.columns],
                'style':'Table Style Medium 9'
            })
            ws.set_column(1,1+nc-1,12,cell_fmt)

            # Hojas por año
            def tipo_ord(p): return 0 if p.upper().startswith('D-') else 1
            def sufijo(p):
                parts = p.upper().split('-')
                return parts[3] if len(parts)>=4 else ''

            for y in years:
                ordered = sorted(
                    data_by_year[y],
                    key=lambda r: (
                        tipo_ord(r['registro_ppu']),
                        parse_ppu(r['registro_ppu'])[0],
                        sufijo(r['registro_ppu'])
                    )
                )
                dfy = pd.DataFrame(ordered).drop(columns=['id','_prohibida'], errors='ignore')
                dfy.to_excel(writer, str(y), startrow=1, startcol=1, index=False, header=True)
                ws2 = writer.sheets[str(y)]
                ws2.merge_range(0,1,0,dfy.shape[1],
                                f"Base de datos del año {y} - {display} a la fecha de {fecha_cod}", title_fmt)
                nr2, nc2 = len(dfy), len(dfy.columns)
                ws2.add_table(1,1,1+nr2,1+nc2-1,{
                    'columns':[{'header':h} for h in dfy.columns],
                    'style':'Table Style Medium 9'
                })
                ws2.set_column(1,1+nc2-1,12,cell_fmt)
                # Sombrear celdas prohibidas dentro de la tabla
                for idx, r in enumerate(ordered):
                    if r['_prohibida']:
                        for col_idx in range(nc2):
                            val = dfy.iat[idx, col_idx]
                            if pd.isna(val):
                                ws2.write(2+idx, 1+col_idx, "", red_fmt)
                            elif isinstance(val, pd.Timestamp):
                                ws2.write(2+idx, 1+col_idx, val.strftime("%d-%m-%Y %H:%M"), red_fmt)
                            else:
                                ws2.write(2+idx, 1+col_idx, val, red_fmt)

        # 12) Enviar respuesta
        excel_data = output.getvalue()
        response = make_response(excel_data)
        response.headers.set("Content-Disposition", f'attachment; filename="{filename}"')
        response.headers.set("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        return response

    except Exception as e:
        app.logger.error("exportar_excel error: %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500

    finally:
        cursor.close()
        conn.close()






@app.route('/api/get_registros', methods=['GET'])
@login_required
def get_registros():
    tipo = request.args.get('tipo')
    year = request.args.get('year')

    connection = get_db_connection()
    if connection is None:
        return jsonify({"error": "Error al conectar con la base de datos"}), 500

    try:
        cursor = connection.cursor(dictionary=True)
        if tipo == 'LEGAJO':
            regex_pattern = r'^(L\. ?|LEG-)(\d{1,4})-{}($|-[A-Z]+$)'.format(year)
            prefix_condition = "registro_ppu LIKE 'L.%' OR registro_ppu LIKE 'LEG-%'"
        elif tipo == 'DENUNCIA':
            regex_pattern = r'^D-(\d+)-{}($|-[A-Z]+$)'.format(year)
            prefix_condition = "registro_ppu LIKE 'D-%'"
        else:
            return jsonify({"error": "Tipo inválido"}), 400

        cursor.execute(f"""
            SELECT registro_ppu FROM datapenal
            WHERE ({prefix_condition}) AND registro_ppu REGEXP %s
            ORDER BY
              CAST(
                CASE
                  WHEN registro_ppu LIKE 'L.%' THEN
                    SUBSTRING_INDEX(registro_ppu, '-', -1)
                  WHEN registro_ppu REGEXP '-[A-Z]$' THEN
                    SUBSTRING_INDEX(SUBSTRING_INDEX(registro_ppu, '-', -2), '-', 1)
                  ELSE
                    SUBSTRING_INDEX(registro_ppu, '-', -1)
                END AS UNSIGNED
              ) ASC,
              CAST(
                CASE
                  WHEN registro_ppu LIKE 'L.%' THEN
                    SUBSTRING(
                      registro_ppu,
                      LOCATE('.', registro_ppu) + 1,
                      LOCATE('-', registro_ppu) - LOCATE('.', registro_ppu) - 1
                    )
                  ELSE
                    SUBSTRING_INDEX(SUBSTRING_INDEX(registro_ppu, '-', 2), '-', -1)
                END AS UNSIGNED
              ) ASC,
              CASE
                WHEN registro_ppu REGEXP '-[A-Z]$' THEN 1
                ELSE 0
              END,
              registro_ppu
        """, (regex_pattern,))

        registros = cursor.fetchall()
        registros_list = [reg['registro_ppu'] for reg in registros]

        return jsonify({"data": registros_list})

    except Exception as e:
        print(f"Error al obtener registros: {e}")
        return jsonify({"error": f"Error al obtener registros: {e}"}), 500
    finally:
        cursor.close()
        connection.close()


@app.route('/api/actualizar_caso', methods=['POST'])
@login_required
def actualizar_caso():
    data = request.json
    registro_ppu = data.get('registro_ppu')
    updated_data = data.get('data')

    if not registro_ppu or not updated_data:
        return jsonify({"error": "Registro PPU y datos a actualizar son requeridos"}), 400

    user_role = session.get('role')

    # Restricción de campos según rol:
    if user_role == 'admin':
        allowed_fields = ['abogado', 'denunciado', 'origen', 'nr de exp completo', 'delito',
                          'departamento', 'fiscalia', 'juzgado', 'informe_juridico', 'item', 'e_situacional', 'etiqueta']
    elif user_role == 'user':
        # Para usuarios, solo se permite modificar "etiqueta"
        allowed_fields = ['etiqueta']
    else:
        return jsonify({"error": "No autorizado"}), 403

    data_to_update = {key: value for key, value in updated_data.items() if key in allowed_fields}

    expediente_juzgado = data_to_update.pop('expediente_juzgado', None)
    if expediente_juzgado:
        if not isinstance(expediente_juzgado, dict):
            return jsonify({"error": "El campo 'expediente_juzgado' debe ser un objeto."}), 400

        errores = validate_expediente_juzgado(expediente_juzgado)
        if errores:
            return jsonify({"error": errores}), 400

        expediente_formateado = (
            f"Exp. {expediente_juzgado['campo1']}-"
            f"{expediente_juzgado['campo2']}-"
            f"{expediente_juzgado['campo3']}-"
            f"{expediente_juzgado['campo4']}-"
            f"{expediente_juzgado['campo5']}-"
            f"{expediente_juzgado['campo6']}-"
            f"{expediente_juzgado['campo7']}"
        )
        existing_origen = data_to_update.get('origen', '').strip()
        if existing_origen:
            existing_origen += f", {expediente_formateado}"
        else:
            existing_origen = expediente_formateado
        data_to_update['origen'] = existing_origen
    else:
        if 'origen' in data_to_update and data_to_update['origen']:
            if data_to_update['origen'][0].isdigit() and not data_to_update['origen'].startswith('CASO '):
                data_to_update['origen'] = 'CASO ' + data_to_update['origen']

    # Se asigna la fecha de última modificación
    data_to_update['last_modified'] = datetime.now()

    connection = get_db_connection()
    if connection is None:
        return jsonify({"error": "Error al conectar con la base de datos"}), 500

    try:
        cursor = connection.cursor()
        # Establecer la variable de sesión @current_user para que el trigger registre el usuario que modifica
        current_user = session.get('username', 'sistema')
        cursor.execute("SET @current_user = %s", (current_user,))

        set_clause = ', '.join(f"`{key}` = %s" for key in data_to_update.keys())
        values = tuple(data_to_update.values()) + (registro_ppu,)

        query = f'UPDATE datapenal SET {set_clause} WHERE registro_ppu = %s'
        cursor.execute(query, values)
        connection.commit()
        return jsonify({"message": "Caso actualizado exitosamente"}), 200
    except Exception as e:
        print(f"Error al actualizar caso: {e}")
        return jsonify({"error": "Error al actualizar caso"}), 500
    finally:
        cursor.close()
        connection.close()



### BUSQUEDA FISCALIAS PARA TABLA MINIMA:
@app.route('/api/search_fiscalias', methods=['GET'])
@login_required
def search_fiscalias():
    query = request.args.get('query', '').strip()

    if not query:
        return jsonify({"data": []})

    query_normalized = normalize_text(query)
    query_terms = query_normalized.split()

    connection = get_db_connection()
    if connection is None:
        return jsonify({"error": "Error al conectar con la base de datos"}), 500

    try:
        cursor = connection.cursor(dictionary=True)
        cursor.execute("SELECT fiscalia, nr_de_exp_completo, departamento FROM dependencias_fiscales_mpfn")
        fiscalias = cursor.fetchall()

        results = []
        for fiscalia in fiscalias:
            fiscalia_name_normalized = normalize_text(fiscalia['fiscalia'])
            # Filtrar solo por los términos de búsqueda, sin palabras clave adicionales
            if all(term in fiscalia_name_normalized for term in query_terms):
                results.append(fiscalia)

        return jsonify({"data": results})

    except Exception as e:
        logger.error(f"Error al obtener fiscalias: {e}")
        return jsonify({"error": f"Error al obtener fiscalias: {e}"}), 500
    finally:
        cursor.close()
        connection.close()

@app.route('/api/historial', methods=['GET'])
@login_required
def obtener_historial():
    registro_ppu = request.args.get('registro_ppu')

    if not registro_ppu:
        return jsonify({"error": "Registro PPU es requerido"}), 400

    connection = get_db_connection()
    if connection is None:
        return jsonify({"error": "Error al conectar con la base de datos"}), 500

    try:
        cursor = connection.cursor(dictionary=True)

        # Esto es tu "versión actual" (sin ruta, pues viene de 'datapenal'):
        query_actual = """
            SELECT 
                'ACTUAL' AS version_id,
                abogado, 
                registro_ppu, 
                denunciado, 
                origen, 
                juzgado, 
                fiscalia, 
                departamento, 
                e_situacional, 
                DATE_FORMAT(last_modified, '%d-%m-%Y') AS fecha_version
            FROM datapenal
            WHERE registro_ppu = %s
            LIMIT 1
        """
        cursor.execute(query_actual, (registro_ppu,))
        version_actual = cursor.fetchone()

        # Esto obtiene TODAS las filas del historial (en datapenal_versioning)...
        query_historial = """
            SELECT 
                version_id, 
                abogado, 
                registro_ppu, 
                denunciado, 
                origen, 
                juzgado, 
                fiscalia, 
                departamento, 
                e_situacional, 
                DATE_FORMAT(fecha_version, '%d-%m-%Y') AS fecha_version, 
                usuario_modificacion,
                ruta
            FROM datapenal_versioning
            WHERE registro_ppu = %s
            ORDER BY fecha_version DESC
        """
        cursor.execute(query_historial, (registro_ppu,))
        historial = cursor.fetchall()

        # -- AÑADE ESTE FILTRO --:
        # Filtra solo las filas que tengan ruta != null y != '' 
        historial_filtrado = [
            row for row in historial
            if row.get('ruta') and row['ruta'].strip() and row['ruta'].strip().upper() != "NULL"
        ]

        return jsonify({
            "version_actual": version_actual,
            # en "historial" devuelves solo las filas con ruta válida:
            "historial": historial_filtrado
        })

    except Exception as e:
        print(f"Error al obtener historial: {e}")
        return jsonify({"error": "Error al obtener historial"}), 500
    finally:
        cursor.close()
        connection.close()


@app.route('/api/obtener_por_ppus', methods=['POST'])
@login_required
def obtener_por_ppus():
    data = request.json
    ppus = data.get('registro_ppu', [])
    if not ppus:
        return jsonify({"data": []})
    connection = get_db_connection()
    if connection is None:
        return jsonify({"error": "Error al conectar con la base de datos"}), 500
    try:
        cursor = connection.cursor(dictionary=True)
        format_strings = ','.join(['%s'] * len(ppus))
        query = f"SELECT * FROM datapenal WHERE registro_ppu IN ({format_strings})"
        cursor.execute(query, tuple(ppus))
        resultados = cursor.fetchall()
        return jsonify({"data": resultados})
    except Exception as e:
        print(f"Error al obtener por ppus: {e}")
        return jsonify({"error": f"Error al obtener por ppus: {e}"}), 500
    finally:
        cursor.close()
        connection.close()

import io
import os
import random
import hashlib

from flask import request, jsonify
from PyPDF2 import PdfReader
import pdfplumber
import pikepdf
from simhash import Simhash

# Asegúrate de haber configurado app.logger en tu Flask app al nivel DEBUG.

ALLOWED_EXTENSIONS = {'pdf'}
HAMMING_THRESHOLD = 3


def read_hash_from_pdf(path):
    """
    Abre el PDF con pikepdf y extrae el metadato '/HashSHA256'
    directamente del Info dictionary.
    Devuelve el valor (str) o None si no existe.
    """
    try:
        import pikepdf
        with pikepdf.Pdf.open(path) as pdf:
            info = pdf.docinfo  # el Info dictionary
            # buscamos con y sin barra
            hash_val = info.get('/HashSHA256') or info.get('HashSHA256')
            if hash_val:
                return str(hash_val)
    except Exception as e:
        app.logger.error("Error leyendo metadato HashSHA256 de %s: %s", path, e)
    return None



def extract_text_pages(data, total_pages, trimmed=False):
    """
    Extrae texto de un PDF en bytes:
     - Si trimmed==False, de todas las páginas.
     - Si trimmed==True y total_pages >50, solo primeras 3 y últimas 3.
    """
    text = ""
    try:
        with pdfplumber.open(io.BytesIO(data)) as pdf:
            if trimmed and total_pages > 50:
                # Primeras 3
                for i in range( min(3, total_pages) ):
                    text += pdf.pages[i].extract_text() or ""
                # Últimas 3
                for i in range(max(0, total_pages - 3), total_pages):
                    text += pdf.pages[i].extract_text() or ""
            else:
                for page in pdf.pages:
                    text += page.extract_text() or ""
    except Exception as e:
        nivel = "trimmed" if (trimmed and total_pages > 50) else ""
        app.logger.error("Error extrayendo texto %s de PDF: %s", nivel, e)
    return text


@app.route('/upload', methods=['POST'])
@login_required
@role_required(['admin'])
def upload_file():
    app.logger.info("=== Iniciando endpoint /upload ===")

    # 1. Recibe todos los PDFs
    files = request.files.getlist('file')
    app.logger.debug("Claves en request.files: %s", list(request.files.keys()))
    app.logger.debug("Listado de archivos recibidos: %s", [f.filename for f in files])
    if not files:
        app.logger.warning("No se recibieron archivos bajo la clave 'file'")
        return jsonify({"error": "No se recibieron archivos"}), 400

    upload_folder = app.config['UPLOAD_FOLDER']
    upload_folder_abs = os.path.abspath(upload_folder)
    app.logger.debug("UPLOAD_FOLDER absoluto: %s", upload_folder_abs)

    # 2. Preprocesa: cuenta páginas y fingerprint completo
    pdf_infos = []
    for f in files:
        filename = f.filename
        app.logger.debug("Procesando archivo: %s", filename)

        if not allowed_file(filename):
            app.logger.warning("Extensión no permitida para: %s", filename)
            continue

        data = f.read()
        app.logger.debug("Leídos %d bytes de %s", len(data), filename)

        # 2.a Conteo de páginas
        try:
            reader = PdfReader(io.BytesIO(data))
            pages = len(reader.pages)
            app.logger.debug("%s → %d páginas", filename, pages)
        except Exception as e:
            pages = 0
            app.logger.error("Error contando páginas de %s: %s", filename, e)

        # 2.b Extracción de texto completo
        full_text = extract_text_pages(data, pages, trimmed=False)
        app.logger.debug("Texto completo extraído de %s (longitud %d)", filename, len(full_text))

        # 2.c Fingerprint Simhash completo
        full_fp = Simhash(full_text.split())
        app.logger.debug("Fingerprint completo de %s: %s", filename, hex(full_fp.value))

        pdf_infos.append({
            'data': data,
            'pages': pages,
            'full_fp': full_fp,
            'name': filename
        })

    # 3. Agrupar por registro PPU
    registro_ppu_re = re.compile(r'(D-\d{1,4}-\d{4}(?:-[A-Z])?|LEG-\d{1,4}-\d{4}(?:-[A-Z])?|L\.?\s?\d{1,4}-\d{4}(?:-[A-Z])?)', re.IGNORECASE)

    grouped = {}
    for info in pdf_infos:
        m = registro_ppu_re.search(info['name'])
        ppu = m.group(0).upper() if m else None
        info['ppu'] = ppu
        grouped.setdefault(ppu, []).append(info)

    # 4. Clustering solo dentro de cada grupo PPU
    clusters = []
    for ppu, infos in grouped.items():
        # Si solo hay uno, lo metemos directamente como un cluster
        if len(infos) == 1:
            clusters.append([infos[0]])
            continue

        local_clusters = []
        for info in infos:
            placed = False
            for cluster in local_clusters:
                ref = cluster[0]
                # decidir si usamos trimmed
                use_trimmed = (info['pages'] > 50) or (ref['pages'] > 50)
                if use_trimmed:
                    t1 = extract_text_pages(info['data'], info['pages'], trimmed=True)
                    t2 = extract_text_pages(ref['data'], ref['pages'], trimmed=True)
                    fp1 = Simhash(t1.split())
                    fp2 = Simhash(t2.split())
                    dist = fp1.distance(fp2)
                    app.logger.debug("Trimmed Hamming %s ↔ %s = %d", info['name'], ref['name'], dist)
                else:
                    dist = info['full_fp'].distance(ref['full_fp'])
                    app.logger.debug("Hamming %s ↔ %s = %d", info['name'], ref['name'], dist)

                if dist <= HAMMING_THRESHOLD:
                    app.logger.info("Agrupado %s con %s (PPU %s)", info['name'], ref['name'], ppu)
                    cluster.append(info)
                    placed = True
                    break

            if not placed:
                local_clusters.append([info])

        clusters.extend(local_clusters)

    app.logger.info("Total clusters: %d", len(clusters))

    # 5. Elegir dentro de cada cluster según número de páginas
    seleccionados = []
    for i, cluster in enumerate(clusters):
        max_pag = max(item['pages'] for item in cluster)
        candidatos = [item for item in cluster if item['pages'] == max_pag]
        elegido = candidatos[0] if len(candidatos) == 1 else random.choice(candidatos)
        app.logger.info("Cluster %d → elegido %s (%d páginas)", i, elegido['name'], elegido['pages'])
        seleccionados.append(elegido)

    # 6. Guardar, leer metadato y precargar fecha_notificación
    resultados = []
    cnx2 = mysql.connector.connect(
        host='localhost', database='monitoreo_descargas_sinoe',
        user='root', password='Manuel22'
    )
    cursor2 = cnx2.cursor()
    for info in seleccionados:
        filename = info['name']
        save_path = os.path.abspath(os.path.join(upload_folder_abs, filename))
        app.logger.debug("Ruta de guardado calculada: %s", save_path)

        if not save_path.startswith(upload_folder_abs):
            app.logger.warning("Ruta inválida (se salta): %s", save_path)
            continue

        # 6.a Guardar el PDF
        try:
            with open(save_path, 'wb') as out:
                out.write(info['data'])
            app.logger.info("Guardado exitoso: %s", filename)
        except Exception as e:
            app.logger.error("Error al guardar %s: %s", filename, e)
            continue

        # 6.b Leer SHA-256 desde el metadato '/HashSHA256'
        hash_sha = read_hash_from_pdf(save_path)
        if hash_sha:
            app.logger.info("Leído SHA desde metadatos de %s: %s", filename, hash_sha)
        else:
            app.logger.warning("Metadato /HashSHA256 no encontrado en %s; omito cálculo", filename)

        # 6.c Buscar fecha_notificación en conteo_exp
        fecha_notif = ""
        if hash_sha:
            cursor2.execute(
                "SELECT fecha_notificacion FROM conteo_exp "
                "WHERE codigo_unico = %s AND fecha_notificacion REGEXP %s",
                (hash_sha, r'^[0-9]{4}-[0-9]{2}-[0-9]{2}$')
            )
            row = cursor2.fetchone()
            if row and row[0]:
                fecha_notif = row[0].isoformat()
                app.logger.debug("Fecha notificación para %s: %s", hash_sha, fecha_notif)

        resultados.append({
            'filename': filename,
            'hash_sha': hash_sha or "",
            'fecha_notificacion': fecha_notif
        })

    cursor2.close()
    cnx2.close()

    app.logger.info("=== Fin de /upload → resultados: %s ===", resultados)
    return jsonify({
        "message": "Carga completada",
        "archivos": resultados
    }), 200
########### PRECARGA FECHA PENAL


@app.route('/api/eliminar_pdfs_por_registro', methods=['POST'])
@login_required
@role_required(['admin'])
def eliminar_pdfs_por_registro():
    data = request.json
    registro_ppu = data.get('registro_ppu', '').strip()
    
    if not registro_ppu:
        return jsonify({"error": "registro_ppu es requerido"}), 400
    
    pdf_folder = app.config['UPLOAD_FOLDER']
    deleted_files = []
    for filename in os.listdir(pdf_folder):
        if filename.lower().endswith('.pdf') and registro_ppu.upper() in filename.upper():
            file_path = os.path.join(pdf_folder, filename)
            if os.path.exists(file_path):
                try:
                    os.remove(file_path)
                    deleted_files.append(filename)
                except Exception as e:
                    return jsonify({"error": f"No se pudo eliminar {filename}: {e}"}), 500
    
    return jsonify({"message": "Archivos eliminados", "deleted": deleted_files}), 200

@app.route('/api/limpiar_pdfs_por_registros', methods=['POST'])
@login_required
@role_required(['admin'])
def limpiar_pdfs_por_registros():
    data = request.json
    ppus = data.get('ppus', [])

    if not ppus:
        return jsonify({"error": "Se requiere una lista de registros PPU"}), 400

    pdf_folder = app.config['UPLOAD_FOLDER']
    deleted_files = []
    for filename in os.listdir(pdf_folder):
        if filename.lower().endswith('.pdf'):
            if any(rppu.upper() in filename.upper() for rppu in ppus):
                file_path = os.path.join(pdf_folder, filename)
                if os.path.exists(file_path):
                    try:
                        os.remove(file_path)
                        deleted_files.append(filename)
                    except Exception as e:
                        return jsonify({"error": f"No se pudo eliminar {filename}: {e}"}), 500

    return jsonify({"message": "Archivos eliminados", "deleted": deleted_files}), 200



@app.route('/api/bulk_update', methods=['POST'])
@login_required
@role_required(['admin'])
def bulk_update():
    """
    Actualiza en masa los registros de la tabla principal según el patrón:
      - Si el registro tiene 'registro_ppu', se actualiza en la tabla "datapenal".
      - Si no tiene 'registro_ppu' pero tiene 'consulta_ppu', se actualiza en la tabla "consulta_ppupenal".
    La lógica de cálculo de hashes, copia y movimiento de PDF se aplica para las tablas
    'datapenal_plazos' y 'datapenal_versioning'. Si en el payload no se envía 'fileName',
    se usará 'ruta', considerándose el registro como nuevo y forzándose la inserción en versioning.
    """

    # ─── Mejora: manejo de multipart/form-data con PDFs ─────────
    if request.content_type.startswith('multipart/form-data'):
        # extraer registros JSON y archivos
        registros = json.loads(request.form.get('registros', '[]'))
        archivos = request.files.getlist('pdfs')
        # guardar cada PDF y asignar fileName al registro correspondiente
        for idx, f in enumerate(archivos):
            filename = secure_filename(f.filename)
            temp_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            f.save(temp_path)
            if idx < len(registros):
                registros[idx]['fileName'] = filename
        # si hay menos archivos que registros, usar el primero para el resto
        if archivos and len(archivos) < len(registros):
            default_name = secure_filename(archivos[0].filename)
            for reg in registros[len(archivos):]:
                reg['fileName'] = default_name
    else:
        # lógica original para JSON puro
        data = request.json
        registros = data.get('registros', [])

    user_role = session.get('role')
    if user_role not in ['admin', 'user']:
        logger.warning("Acceso no autorizado intentado.")
        return jsonify({"error": "No autorizado"}), 403

    # Preparar carpeta destino según fecha
    months_map = {
        1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL",
        5: "MAYO", 6: "JUNIO", 7: "JULIO", 8: "AGOSTO",
        9: "SETIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"
    }
    today = datetime.now()
    year_str = str(today.year)
    month_str = months_map.get(today.month, "MES_DESCONOCIDO")
    day_str = today.strftime("%d-%m-%Y")
    final_folder = os.path.join(DESTINATION_BASE_PATH, year_str, month_str, day_str)
    os.makedirs(final_folder, exist_ok=True)
    logger.info(f"Carpeta destino preparada: {final_folder}")

    def sha256_of_file(filepath):
        """Calcula el hash SHA-256 de un archivo."""
        sha256 = hashlib.sha256()
        try:
            with open(filepath, 'rb') as f:
                for chunk in iter(lambda: f.read(4096), b""):
                    sha256.update(chunk)
            return sha256.hexdigest()
        except Exception as e:
            logger.error(f"Error al calcular SHA-256 del archivo {filepath}: {e}")
            return None

    connection_registro = None
    connection_consulta = None
    cursor_registro = None
    cursor_consulta = None
    # Lista para almacenar los archivos PDF temporales que se eliminarán al final
    pdfs_to_delete = []

    try:
        # Conectar a la BD (la misma base para ambos casos)
        connection_registro = get_db_connection()
        connection_consulta = get_db_connection()
        if connection_registro is None or connection_consulta is None:
            logger.error("Error al conectar con la base de datos.")
            return jsonify({"error": "Error al conectar con la base de datos"}), 500

        cursor_registro = connection_registro.cursor(dictionary=True)
        cursor_consulta = connection_consulta.cursor(dictionary=True)
        logger.info("Conexión a la base de datos establecida correctamente.")

        for reg in registros:
            logger.info(f"Iniciando procesamiento del registro: {reg}")
            # Se obtiene el identificador: se usa 'registro_ppu' o, si está vacío, 'consulta_ppu'
            accion_val = (reg.get('accion', '') or '').strip()
            registro_ppu = (reg.get('registro_ppu', '') or '').strip()
            if not registro_ppu:
                registro_ppu = (reg.get('consulta_ppu', '') or '').strip()

            # Se obtiene el nombre del archivo; si falta 'fileName', se usa 'ruta'
            original_file_name = (reg.get('fileName') or reg.get('ruta') or '').strip()
            if not registro_ppu or not original_file_name:
                logger.warning(f"Registro omitido por falta de identificador o nombre de archivo: {reg}")
                continue

            pdf_path = os.path.join(app.config['UPLOAD_FOLDER'], original_file_name)
            if not os.path.isfile(pdf_path):
                logger.warning(f"Archivo no encontrado: {pdf_path}")
                continue

            # Agregar el PDF a la lista para eliminarlo al finalizar
            pdfs_to_delete.append(pdf_path)
            logger.info(f"PDF a procesar: {pdf_path}")

            # Determinar el tipo: si se especifica "tipo" o se infiere que es consulta (si existe consulta_ppu)
            tipo = (reg.get('tipo') or '').strip().lower()
            if not tipo and reg.get('consulta_ppu'):
                tipo = "consulta_ppu"

            # Bifurcación: buscar en la tabla correcta
            if tipo == 'consulta_ppu':
                main_table = "consulta_ppupenal"
                cur = cursor_consulta
                cur.execute("SELECT * FROM consulta_ppupenal WHERE consulta_ppu = %s LIMIT 1", (registro_ppu,))
            else:
                main_table = "datapenal"
                cur = cursor_registro
                cur.execute("SELECT * FROM datapenal WHERE registro_ppu = %s LIMIT 1", (registro_ppu,))
            row_dp = cur.fetchone()
            if not row_dp:
                logger.warning(f"No se encontró registro en '{main_table}' para {registro_ppu}")
                continue

            # Si no se envía fileName, se asume que se usó "ruta" y se considera registro nuevo.
            registro_nuevo = not reg.get('fileName')

            # Detección de cambio en e_situacional:
            # Se fuerza el cambio si el campo almacenado es nulo, vacío, "ingreso nuevo" o si el registro es nuevo.
            current_e = row_dp.get('e_situacional')
            if registro_nuevo or current_e in [None, ''] or (isinstance(current_e, str) and current_e.strip().lower() == "ingreso nuevo"):
                situacion_cambio = True
            else:
                old_situacion = current_e.strip().lower()
                new_situacion = reg.get('e_situacional', current_e).strip().lower()
                situacion_cambio = (new_situacion != old_situacion)
            logger.info(f"Registro {registro_ppu}: situacion_cambio = {situacion_cambio}")

            # Consultar el último registro en datapenal_plazos para ver si la acción cambió
            cur.execute("""
                SELECT accion
                FROM datapenal_plazos
                WHERE registro_ppu = %s
                ORDER BY last_modified DESC
                LIMIT 1
            """, (registro_ppu,))
            last_plazos = cur.fetchone()
            accion_actual = (last_plazos['accion'].strip() if last_plazos and last_plazos.get('accion') else '')
            accion_modificada = (accion_val != accion_actual) and (accion_val != '')
            logger.info(f"Registro {registro_ppu}: accion_actual = '{accion_actual}', accion_modificada = {accion_modificada}")

            # Actualizar fiscalía y departamento
            fiscalia_actualizada = reg.get('fiscalia', row_dp.get('fiscalia', ''))
            cur.execute("SELECT departamento FROM dependencias_fiscales_mpfn WHERE fiscalia = %s LIMIT 1", (fiscalia_actualizada,))
            fiscalia_row = cur.fetchone()
            if not fiscalia_row:
                departamento_actualizado = row_dp.get('departamento', '')
                fiscalia_actualizada = row_dp.get('fiscalia', '')
                logger.warning(f"La fiscalía '{fiscalia_actualizada}' no existe en la BD.")
            else:
                departamento_actualizado = fiscalia_row['departamento']
            logger.info(f"Registro {registro_ppu}: fiscalia = {fiscalia_actualizada}, departamento = {departamento_actualizado}")

            # Procesar fecha de ingreso
            fecha_ingreso_val = reg.get('fecha_ingreso', row_dp.get('fecha_ingreso', None))
            if isinstance(fecha_ingreso_val, date):
                fecha_ingreso_str = fecha_ingreso_val.strftime("%Y-%m-%d")
            elif fecha_ingreso_val:
                try:
                    parsed_fecha_ingreso = parser.parse(fecha_ingreso_val)
                    fecha_ingreso_str = parsed_fecha_ingreso.strftime("%Y-%m-%d")
                except Exception as e:
                    logger.error(f"Error al parsear 'fecha_ingreso' para {registro_ppu}: {e}")
                    fecha_ingreso_str = None
            else:
                fecha_ingreso_str = None

            # Armar datos de actualización para la tabla principal
            update_data = {
                'abogado': reg.get('abogado', row_dp.get('abogado', '')),
                'denunciado': reg.get('denunciado', row_dp.get('denunciado', '')),
                'origen': reg.get('origen', row_dp.get('origen', '')),
                'nr de exp completo': reg.get('nr de exp completo', row_dp.get('nr de exp completo', '')),
                'delito': reg.get('delito', row_dp.get('delito', '')),
                'departamento': departamento_actualizado,
                'fiscalia': fiscalia_actualizada,
                'juzgado': reg.get('juzgado', row_dp.get('juzgado', '')),
                'informe_juridico': reg.get('informe_juridico', row_dp.get('informe_juridico', '')),
                'item': reg.get('item', row_dp.get('item', '')),
                'e_situacional': reg.get('e_situacional', row_dp.get('e_situacional', '')),
                'fecha_ingreso': fecha_ingreso_str,
                'last_modified': datetime.now(),
                'etiqueta': reg.get('etiqueta', row_dp.get('etiqueta', ''))
            }
            if tipo == 'consulta_ppu':
                update_data['consulta_ppu'] = reg.get('consulta_ppu', 'consulta_ppu')

            set_clause = ', '.join(f"`{k}` = %s" for k in update_data.keys())
            val_dp = tuple(update_data.values()) + (registro_ppu,)
            sql_update = f"UPDATE {main_table} SET {set_clause} WHERE registro_ppu = %s"
            try:
                cur.execute(sql_update, val_dp)
                logger.info(f"Actualizado '{main_table}' para {registro_ppu}")
            except mysql.connector.Error as err:
                logger.error(f"Error al actualizar '{main_table}' para {registro_ppu}: {err}")
                continue

            # Procesar campos de audiencia y plazo_atencion
            audiencia_val = reg.get('audiencia', False)
            plazo_atencion_val = reg.get('plazo_atencion', '').strip()
            if audiencia_val:
                if not plazo_atencion_val:
                    logger.warning(f"'plazo_atencion' requerido para audiencia en {registro_ppu}")
                    continue
                try:
                    plazo_atencion_val = re.sub(r"(?<=-)(0)(\d)(?=-)", r"\2", plazo_atencion_val)
                    try:
                        parsed_dt = datetime.strptime(plazo_atencion_val, "%d-%m-%Y %I:%M %p")
                    except ValueError:
                        parsed_dt = datetime.strptime(plazo_atencion_val, "%d-%m-%Y %H:%M %p")
                    plazo_atencion_formatted = parsed_dt.strftime("%d-%m-%Y %H:%M")
                except ValueError:
                    logger.warning(f"Formato inválido de 'plazo_atencion' para audiencia en {registro_ppu}")
                    continue
            else:
                if plazo_atencion_val and not plazo_atencion_val.isdigit():
                    logger.warning(f"'plazo_atencion' debe ser numérico para {registro_ppu}")
                    continue
                plazo_atencion_formatted = plazo_atencion_val

            # Procesar fecha_atencion
            fecha_atencion_val = reg.get('fecha_atencion', '').strip()
            if fecha_atencion_val:
                try:
                    parsed_fecha_atencion = parser.parse(fecha_atencion_val)
                    fecha_atencion_str = parsed_fecha_atencion.strftime("%Y-%m-%d")
                except Exception as e:
                    logger.error(f"Error al parsear 'fecha_atencion' para {registro_ppu}: {e}")
                    fecha_atencion_str = None
            else:
                fecha_atencion_str = None

            # Decidir mover PDF: si hay cambio en e_situacional, en acción o si la ruta almacenada es incompleta (no empieza con "\\")
            mover_pdf = (situacion_cambio or accion_modificada)
            if not row_dp.get('ruta', '').startswith('\\\\'):
                mover_pdf = True

            # Calcular hash SHA-256 del archivo PDF
            hash_sha = sha256_of_file(pdf_path)
            if not hash_sha:
                logger.error(f"No se pudo calcular el hash para {pdf_path}")
                continue

            # Consultar si ya existe el hash en datapenal_plazos y versioning
            cur.execute("SELECT ruta FROM datapenal_plazos WHERE hash_sha = %s LIMIT 1", (hash_sha,))
            plazos_row_con_hash = cur.fetchone()
            cur.execute("SELECT ruta FROM datapenal_versioning WHERE hash_sha = %s LIMIT 1", (hash_sha,))
            versioning_row_con_hash = cur.fetchone()

            if mover_pdf:
                if not plazos_row_con_hash and not versioning_row_con_hash:
                    # Mantener el mismo nombre base; si existe, usar " (1)", " (2)", …
                    def ensure_unique_name(folder: str, fname: str) -> str:
                        p = Path(folder) / fname
                        if not p.exists():
                            return str(p)
                        stem, ext = p.stem, p.suffix
                        i = 1
                        candidate = Path(folder) / f"{stem} ({i}){ext}"
                        while candidate.exists():
                            i += 1
                            candidate = Path(folder) / f"{stem} ({i}){ext}"
                        return str(candidate)

                    dest_path = ensure_unique_name(final_folder, original_file_name)

                    try:
                        tmp_path = f"{dest_path}.part"
                        with open(pdf_path, 'rb') as src, open(tmp_path, 'wb') as dst:
                            shutil.copyfileobj(src, dst, length=1048576)
                        os.replace(tmp_path, dest_path)  # atómico si es la misma unidad
                        logger.info(f"PDF copiado a: {dest_path}")
                    except Exception as e:
                        try:
                            if os.path.exists(tmp_path):
                                os.remove(tmp_path)
                        except Exception:
                            pass
                        logger.error(f"Error al copiar {pdf_path} a {dest_path}: {e}")
                        continue

                    ruta_completa = os.path.normpath(dest_path)
                else:
                    if plazos_row_con_hash:
                        ruta_completa = plazos_row_con_hash['ruta']
                    else:
                        ruta_completa = versioning_row_con_hash['ruta']
                    logger.info(f"PDF no se copia. Se reutiliza la ruta: {ruta_completa}")


                # Insertar en datapenal_plazos si la acción ha cambiado
                if accion_modificada:
                    cur.execute("""
                        SELECT plazo_atencion
                        FROM datapenal_plazos
                        WHERE hash_sha = %s AND registro_ppu = %s AND tipoPlazo = %s
                    """, (hash_sha, registro_ppu, reg.get('tipoPlazo', '')))
                    filas_mismo_hash = cur.fetchall()
                    if filas_mismo_hash:
                        plazos_existentes = [fila['plazo_atencion'] for fila in filas_mismo_hash]
                        if plazo_atencion_formatted in plazos_existentes:
                            logger.warning(f"Hash y plazo_atencion idénticos en datapenal_plazos para {registro_ppu} y tipoPlazo {reg.get('tipoPlazo')}. Se omite inserción.")
                        else:
                            insert_data_plazos = {
                                'abogado': row_dp.get('abogado', ''),
                                'registro_ppu': registro_ppu,
                                'denunciado': row_dp.get('denunciado', ''),
                                'origen': row_dp.get('origen', ''),
                                'nr de exp completo': row_dp.get('nr de exp completo', ''),
                                'delito': row_dp.get('delito', ''),
                                'departamento': departamento_actualizado,
                                'fiscalia': fiscalia_actualizada,
                                'juzgado': row_dp.get('juzgado', ''),
                                'informe_juridico': row_dp.get('informe_juridico', ''),
                                'item': row_dp.get('item', ''),
                                'e_situacional': reg.get('e_situacional', row_dp.get('e_situacional', '')),
                                'fecha_ingreso': fecha_ingreso_str if fecha_ingreso_str else row_dp.get('fecha_ingreso', None),
                                'last_modified': datetime.now(),
                                'etiqueta': row_dp.get('etiqueta', ''),
                                'accion': accion_val,
                                'plazo_atencion': plazo_atencion_formatted,
                                'fecha_atencion': fecha_atencion_str,
                                'seguimiento': None,
                                'ruta': ruta_completa,
                                'hash_sha': hash_sha,
                                'tipoPlazo': reg.get('tipoPlazo', '')
                            }
                            if tipo == 'consulta_ppu':
                                insert_data_plazos['consulta_ppu'] = reg.get('consulta_ppu', 'consulta_ppu')
                            cols_plazos = ', '.join(f"`{k}`" for k in insert_data_plazos.keys())
                            place_plazos = ', '.join(['%s'] * len(insert_data_plazos))
                            vals_plazos = tuple(insert_data_plazos.values())
                            sql_insert_plazos = f"INSERT INTO datapenal_plazos ({cols_plazos}) VALUES ({place_plazos})"
                            try:
                                cur.execute(sql_insert_plazos, vals_plazos)
                                logger.info(f"Insertado en datapenal_plazos para {registro_ppu} con plazo diferente.")
                            except mysql.connector.Error as err:
                                logger.error(f"Error en INSERT datapenal_plazos para {registro_ppu}: {err}")
                    else:
                        insert_data_plazos = {
                            'abogado': row_dp.get('abogado', ''),
                            'registro_ppu': registro_ppu,
                            'denunciado': row_dp.get('denunciado', ''),
                            'origen': row_dp.get('origen', ''),
                            'nr de exp completo': row_dp.get('nr de exp completo', ''),
                            'delito': row_dp.get('delito', ''),
                            'departamento': departamento_actualizado,
                            'fiscalia': fiscalia_actualizada,
                            'juzgado': row_dp.get('juzgado', ''),
                            'informe_juridico': row_dp.get('informe_juridico', ''),
                            'item': row_dp.get('item', ''),
                            'e_situacional': reg.get('e_situacional', row_dp.get('e_situacional', '')),
                            'fecha_ingreso': fecha_ingreso_str if fecha_ingreso_str else row_dp.get('fecha_ingreso', None),
                            'last_modified': datetime.now(),
                            'etiqueta': row_dp.get('etiqueta', ''),
                            'accion': accion_val,
                            'plazo_atencion': plazo_atencion_formatted,
                            'fecha_atencion': fecha_atencion_str,
                            'seguimiento': None,
                            'ruta': ruta_completa,
                            'hash_sha':	hash_sha,
                            'tipoPlazo': reg.get('tipoPlazo', '')
                        }
                        if tipo == 'consulta_ppu':
                            insert_data_plazos['consulta_ppu'] = reg.get('consulta_ppu', 'consulta_ppu')
                        cols_plazos = ', '.join(f"`{k}`" for k in insert_data_plazos.keys())
                        place_plazos = ', '.join(['%s'] * len(insert_data_plazos))
                        vals_plazos = tuple(insert_data_plazos.values())
                        sql_insert_plazos = f"INSERT INTO datapenal_plazos ({cols_plazos}) VALUES ({place_plazos})"
                        try:
                            cur.execute(sql_insert_plazos, vals_plazos)
                            logger.info(f"Insertado en datapenal_plazos para {registro_ppu}")
                        except mysql.connector.Error as err:
                            logger.error(f"Error en INSERT datapenal_plazos para {registro_ppu}: {err}")

                # Inserción en versioning: se inserta si se detecta un cambio en e_situacional
                # o si no existe registro previo para el hash
                cur.execute("SELECT COUNT(*) AS c FROM datapenal_versioning WHERE hash_sha = %s", (hash_sha,))
                count_row = cur.fetchone()
                if situacion_cambio or (count_row and count_row.get('c', 0) == 0):
                    insert_data_versioning = {
                        'abogado': row_dp.get('abogado', ''),
                        'registro_ppu': registro_ppu,
                        'denunciado': row_dp.get('denunciado', ''),
                        'origen': row_dp.get('origen', ''),
                        'fiscalia': row_dp.get('fiscalia', ''),
                        'juzgado': row_dp.get('juzgado', ''),
                        'departamento': row_dp.get('departamento', ''),
                        'e_situacional': reg.get('e_situacional', row_dp.get('e_situacional', '')),
                        'fecha_version': datetime.now(),
                        'usuario_modificacion': session.get('username', 'desconocido'),
                        'ruta': ruta_completa,
                        'hash_sha': hash_sha
                    }
                    if tipo == 'consulta_ppu':
                        insert_data_versioning['consulta_ppu'] = reg.get('consulta_ppu', 'consulta_ppu')
                    cols_ver = ', '.join(f"`{k}`" for k in insert_data_versioning.keys())
                    place_ver = ', '.join(['%s'] * len(insert_data_versioning))
                    vals_ver = tuple(insert_data_versioning.values())
                    sql_insert_ver = f"INSERT INTO datapenal_versioning ({cols_ver}) VALUES ({place_ver})"
                    try:
                        cur.execute(sql_insert_ver, vals_ver)
                        logger.info(f"Insertado en datapenal_versioning para {registro_ppu}")
                    except mysql.connector.Error as err:
                        logger.error(f"Error en INSERT datapenal_versioning para {registro_ppu}: {err}")

        # Listar todos los PDFs procesados
        logger.info("Listado de todos los PDFs procesados:")
        for pdf in pdfs_to_delete:
            logger.info(f"PDF procesado: {pdf}")

        # Eliminar los archivos PDF temporales al finalizar
        logger.info("Eliminando archivos PDF temporales...")
        for pdf in pdfs_to_delete:
            try:
                os.remove(pdf)
                logger.info(f"PDF temporal eliminado: {pdf}")
            except Exception as e:
                logger.error(f"Error al eliminar {pdf}: {e}")

        connection_registro.commit()
        connection_consulta.commit()
        logger.info("Bulk update completado exitosamente.")
        return jsonify({"message": "Actualización masiva completada"}), 200

    except Exception as e:
        logger.error(f"Error en bulk_update: {e}", exc_info=True)
        if connection_registro:
            connection_registro.rollback()
        if connection_consulta:
            connection_consulta.rollback()
        return jsonify({"error": f"Error en bulk_update: {e}"}), 500

    finally:
        try:
            if cursor_registro:
                cursor_registro.close()
            if connection_registro:
                connection_registro.close()
            if cursor_consulta:
                cursor_consulta.close()
            if connection_consulta:
                connection_consulta.close()
        except Exception as e:
            logger.error(f"Error al cerrar conexión o cursor: {e}")



## Agregar que 'e_situacional' pase con el nuevo valor

@app.route('/api/datapenal_plazos', methods=['GET'])
@login_required
def get_datapenal_plazos():
    """
    Endpoint que devuelve los registros de 'datapenal_plazos' con los filtros:
      - query (busca en múltiples columnas de datapenal y datapenal_plazos)
      - page (paginación)
      - limit (cantidad de filas por página)
      - abogado (filtra solo las filas cuyo dp.abogado contenga el valor)
      - mostrar_archivados (booleano para excluir registros con etiqueta='ARCHIVO')
      
    También fuerza el filtro de abogado si el role del usuario es 'user'.
    Incluye el campo observacion_abogado en la respuesta JSON.
    """
    query_param = request.args.get('query', '').strip()
    page = int(request.args.get('page', 1))
    limit = int(request.args.get('limit', 200))
    offset = (page - 1) * limit

    # ===========================
    # Determinar filtro por abogado según rol
    user_role = session.get('role')
    current_username = session.get('username', '')
    if user_role == 'user':
        # Si es 'user', forzamos el abogado según el mapeo
        abogado_filter = username_to_abogado.get(current_username, '').upper()
    else:
        # Si es admin/coordinador, tomamos el parámetro 'abogado'
        abogado_filter = request.args.get('abogado', '')
        if ';' in abogado_filter:
            abogado_filter = abogado_filter.split(';')[-1]
        abogado_filter = abogado_filter.strip().upper()
    # ===========================

    mostrar_archivados = request.args.get('mostrar_archivados', 'true').lower() == 'true'

    connection = get_db_connection()
    if connection is None:
        return jsonify({"error": "Error al conectar con la base de datos"}), 500

    try:
        conditions = []
        params = []
        cursor = connection.cursor(dictionary=True)

        if not mostrar_archivados:
            conditions.append("(dp.etiqueta IS NULL OR dp.etiqueta != %s)")
            params.append('ARCHIVO')

        if abogado_filter:
            # Filtrar por abogado en dp.abogado (parte posterior al “;” si existe)
            conditions.append("""
                CASE
                  WHEN LOCATE(';', dp.abogado) > 0
                    THEN TRIM(SUBSTRING_INDEX(dp.abogado, ';', -1))
                  ELSE
                    TRIM(dp.abogado)
                END = %s
            """)
            params.append(abogado_filter.lower())

        if query_param:
            # Columnas en las que buscar el texto de 'query'
            columns = [
                'dp.abogado',
                'dp.denunciado',
                'dp.origen',
                'dp.delito',
                'dp.departamento',
                'dp.fiscalia',
                'dp.juzgado',
                'dp_plazos.registro_ppu',
                'dp_plazos.e_situacional'
            ]
            query_clauses = " OR ".join([f"LOWER({col}) LIKE LOWER(%s)" for col in columns])
            conditions.append(f"({query_clauses})")
            params.extend([f"%{query_param}%"] * len(columns))

        where_clause = f"WHERE {' AND '.join(conditions)}" if conditions else ""

        # 1. Contar total de registros
        count_query = f"""
            SELECT COUNT(*) AS total
            FROM datapenal_plazos dp_plazos
            JOIN datapenal dp ON dp_plazos.registro_ppu = dp.registro_ppu
            {where_clause}
        """
        cursor.execute(count_query, params)
        total_records = cursor.fetchone()['total']

        # 2. Obtener datos paginados, incluyendo 'observacion_abogado',
        #    y dando prioridad a seguimiento='OBSERVADO' (grupo primero), luego por registro_ppu
        data_query = f"""
            SELECT 
                dp_plazos.id,
                dp_plazos.e_situacional,
                dp.abogado,
                dp_plazos.accion,
                dp_plazos.plazo_atencion,
                dp_plazos.seguimiento,
                dp_plazos.ruta,
                dp_plazos.fecha_atencion,
                dp_plazos.registro_ppu,
                dp.denunciado,
                dp.origen,
                dp.fiscalia,
                dp.juzgado,
                dp.departamento,
                dp_plazos.observacion_abogado
            FROM datapenal_plazos dp_plazos
            JOIN datapenal dp ON dp_plazos.registro_ppu = dp.registro_ppu
            {where_clause}
            ORDER BY
                -- 1) Filas con seguimiento = 'OBSERVADO' primero (devuelven 1), resto 0
                CASE WHEN dp_plazos.seguimiento = 'OBSERVADO' THEN 1 ELSE 0 END DESC,
                -- 2) Dentro de cada grupo, por registro_ppu (ajusta aquí si tienes otro campo de urgencia)
                dp_plazos.registro_ppu ASC
            LIMIT %s OFFSET %s
        """
        cursor.execute(data_query, params + [limit, offset])
        resultados = cursor.fetchall()

        # Formatear 'fecha_atencion' como cadena legible
        for row in resultados:
            if isinstance(row.get('fecha_atencion'), datetime):
                row['fecha_atencion'] = row['fecha_atencion'].strftime('%Y-%m-%d %H:%M:%S')
            # Si no existe, deja el valor tal cual (NULL o texto)

            # Asegurarse de que observacion_abogado sea al menos cadena vacía
            row['observacion_abogado'] = row.get('observacion_abogado') or ""

        total_pages = (total_records + limit - 1) // limit

        return jsonify({
            "data": resultados,
            "page": page,
            "total_pages": total_pages,
            "total_records": total_records
        }), 200

    except Exception as e:
        print(f"Error al obtener datapenal_plazos: {e}")
        return jsonify({"error": "Error al obtener datos"}), 500

    finally:
        if connection.is_connected():
            connection.close()

@app.route('/api/exportar_excel_plazos', methods=['GET'])
@login_required
def exportar_excel_plazos():
    """
    Exporta a Excel los datos de la tabla datapenal_plazos (JOIN datapenal),
    aplicando SOLO los filtros básicos:
      - abogado (forzado si es user)
      - query
      - mostrar_archivados
    y SIN excluir filas por 'seguimiento=ATENDIDA', NI lógica de 'vencido'.
    """
    query = request.args.get('query', '').strip()
    mostrar_archivados = request.args.get('mostrar_archivados', 'true').lower() == 'true'

    # Forzar abogado si es "user"
    user_role = session.get('role')
    current_username = session.get('username', '')
    if user_role == 'user':
        abogado_filter = username_to_abogado.get(current_username, '').upper()
    else:
        # admin => usa param ?abogado=...
        abogado_filter = request.args.get('abogado', '').strip().upper()

    connection = get_db_connection()
    if connection is None:
        return jsonify({"error": "Error al conectar con la base de datos"}), 500

    try:
        cursor = connection.cursor(dictionary=True)

        # 1) Construir condiciones
        conditions = []
        params = []

        # mostrar_archivados => filtrar etiqueta != 'ARCHIVO'
        if not mostrar_archivados:
            conditions.append("(dp.etiqueta IS NULL OR dp.etiqueta != %s)")
            params.append('ARCHIVO')

        # Filtro por abogado
        if abogado_filter:
            conditions.append("LOWER(dp.abogado) LIKE LOWER(%s)")
            params.append(f"%{abogado_filter}%")

        # Filtro por query
        if query:
            # Aplicas el REGEXP o LIKE en las columnas que gustes
            columns = [
                'dp.abogado', 'dp.denunciado', 'dp.origen', 'dp.delito',
                'dp.departamento', 'dp.fiscalia', 'dp.juzgado',
                'dp_plazos.registro_ppu', 'dp_plazos.e_situacional'
            ]
            subconds = []
            for col in columns:
                subconds.append(f"LOWER({col}) LIKE LOWER(%s)")
                params.append(f"%{query}%")
            conditions.append("(" + " OR ".join(subconds) + ")")

        # Generar where_clause
        where_clause = "WHERE " + " AND ".join(conditions) if conditions else ""

        # 2) Query para obtener TODOS los datos (sin paginación)
        #    uniendo datapenal_plazos (dp_plazos) con datapenal (dp)
        sql = f"""
        SELECT
            dp_plazos.id,
            dp_plazos.registro_ppu,
            dp_plazos.fecha_atencion,
            dp_plazos.plazo_atencion,
            dp_plazos.accion,
            dp_plazos.seguimiento,
            dp_plazos.ruta,
            dp_plazos.hash_sha,
            dp_plazos.fecha_ingreso,
            dp_plazos.last_modified,

            dp.abogado,
            dp.denunciado,
            dp.origen,
            dp.delito,
            dp.departamento,
            dp.fiscalia,
            dp.juzgado,
            dp_plazos.e_situacional,
            dp.etiqueta
        FROM datapenal_plazos dp_plazos
        JOIN datapenal dp
          ON dp_plazos.registro_ppu = dp.registro_ppu
        {where_clause}
        """

        cursor.execute(sql, params)
        filas = cursor.fetchall()

        # 3) Opcionalmente, ordenar por parse_ppu (si deseas un orden manual)
        filas_sorted = sorted(filas, key=lambda r: parse_ppu(r['registro_ppu']))

        # 4) Crear DataFrame
        df = pd.DataFrame(filas_sorted)
        # (Opcional) remove columnas innecesarias, p.ej. 'id'
        if not df.empty:
            df.drop(columns=['id'], inplace=True, errors='ignore')

        # 5) Generar Excel en memoria
        output = BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            df.to_excel(writer, index=False, sheet_name='Plazos')
            writer.close()
        output.seek(0)

        # 6) Retornar el Excel
        fecha_actual = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre_archivo = f"plazos_exportados_{fecha_actual}.xlsx"
        return send_file(
            output,
            download_name=nombre_archivo,
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        logger.error(f"Error al exportar Excel plazos: {e}")
        return jsonify({"error": "Error al exportar Excel"}), 500
    finally:
        cursor.close()
        connection.close()



@app.route('/api/descargar_pdf', methods=['GET'])
@login_required
def descargar_pdf():
    ruta = request.args.get('ruta')
    if not ruta:
        return jsonify({"error": "Ruta es requerida"}), 400

    # Definir rutas base permitidas
    ALLOWED_BASE_PATHS = [
        os.path.abspath(app.config['UPLOAD_FOLDER']),
        os.path.abspath(DESTINATION_BASE_PATH)
    ]

    # Convertir la ruta proporcionada a una ruta absoluta
    ruta_abs = os.path.abspath(ruta)

    # Verificar si la ruta absoluta comienza con alguna de las rutas base permitidas
    if not any(ruta_abs.startswith(base_path) for base_path in ALLOWED_BASE_PATHS):
        return jsonify({"error": "Ruta inválida"}), 400

    directory = os.path.dirname(ruta_abs)
    filename = os.path.basename(ruta_abs)

    try:
        return send_from_directory(directory, filename, as_attachment=True)
    except FileNotFoundError:
        return jsonify({"error": "Archivo no encontrado"}), 404

def calcular_fecha_limite(fecha_atencion, plazo_atencion):
    """
    Calcula la fecha límite sumando días hábiles a una fecha de inicio.

    Args:
        fecha_atencion (str): Fecha de inicio en formato "YYYY-MM-DD HH:MM:SS".
        plazo_atencion (int): Número de días hábiles a sumar.

    Returns:
        datetime: Fecha límite.
    """
    fecha_atencion_dt = datetime.strptime(fecha_atencion, "%Y-%m-%d %H:%M:%S")
    fecha_limite = fecha_atencion_dt
    while plazo_atencion > 0:
        fecha_limite += timedelta(days=1)
        if fecha_limite.weekday() < 5:  # 0=Lunes, ..., 4=Viernes
            plazo_atencion -= 1
    return fecha_limite

def calcular_dias_restantes(fecha_limite):
    """
    Calcula los días hábiles restantes entre la fecha actual y la fecha límite.

    Args:
        fecha_limite (datetime): Fecha límite.

    Returns:
        int: Días hábiles restantes. Devuelve -1 si ya venció.
    """
    fecha_hoy = datetime.now()
    delta_dias = (fecha_limite - fecha_hoy).days
    dias_restantes = 0

    for i in range(delta_dias + 1):  # Incluye el día límite
        dia_actual = fecha_hoy + timedelta(days=i)
        if dia_actual.weekday() < 5:  # Excluye fines de semana
            dias_restantes += 1

    return dias_restantes if fecha_limite >= fecha_hoy else -1

@app.route('/api/get_plazos', methods=['GET'])
@login_required
def get_plazos():
    connection = get_db_connection()
    if connection is None:
        return jsonify({"error": "Error al conectar con la base de datos"}), 500

    try:
        cursor = connection.cursor(dictionary=True)
        cursor.execute("""
            SELECT 
                dp_plazos.id,
                dp_plazos.e_situacional,
                dp.abogado,
                dp_plazos.accion,
                dp_plazos.plazo_atencion,
                dp_plazos.seguimiento,
                dp_plazos.ruta,
                dp_plazos.fecha_atencion,
                dp_plazos.registro_ppu,
                dp.denunciado,
                dp.origen,
                dp.fiscalia,
                dp.juzgado,
                dp.departamento
            FROM datapenal_plazos dp_plazos
            JOIN datapenal dp ON dp_plazos.registro_ppu = dp.registro_ppu
        """)
        resultados = cursor.fetchall()

        # Calcular 'fecha_limite' y 'dias_restantes' para cada registro
        plazos_resultados = []
        for row in resultados:
            fecha_atencion = row.get('fecha_atencion')
            plazo_atencion = row.get('plazo_atencion')

            if fecha_atencion and plazo_atencion:
                try:
                    fecha_limite = calcular_fecha_limite(fecha_atencion, plazo_atencion)
                    dias_restantes = calcular_dias_restantes(fecha_limite)
                    row['fecha_limite'] = fecha_limite.strftime('%Y-%m-%d')
                    row['dias_restantes'] = (
                        "Vencido" if dias_restantes < 0 else
                        "URGENTE RESOLVER EN EL DÍA" if dias_restantes == 1 else
                        f"{dias_restantes} días restantes"
                    )
                except Exception as e:
                    row['fecha_limite'] = None
                    row['dias_restantes'] = 'Error en cálculo'
            else:
                row['fecha_limite'] = None
                row['dias_restantes'] = 'Datos incompletos'

            plazos_resultados.append(row)

        return jsonify({"data": plazos_resultados}), 200
    except Exception as e:
        logger.error(f"Error al obtener plazos: {e}")
        return jsonify({"error": "Error al obtener plazos"}), 500
    finally:
        cursor.close()
        connection.close()


@app.route('/api/actualizar_seguimiento', methods=['POST'])
def actualizar_seguimiento():
    logger.info("=== Iniciando proceso de actualización de seguimiento ===")
    row_id = request.form.get('id', '').strip()
    atendido_str = request.form.get('atendido', 'false').lower()
    
    # Loguear las claves enviadas en request.files para depuración
    files_keys = list(request.files.keys())
    logger.debug(f"Claves recibidas en request.files: {files_keys}")
    file = request.files.get('file')

    logger.debug(f"Parametros recibidos: id={row_id}, atendido={atendido_str}, file={'None' if not file else file.filename}")

    if not row_id:
        logger.error("Falta el parámetro 'id'.")
        return jsonify({"error": "id (en datapenal_plazos) es requerido"}), 400

    if atendido_str == 'true' and (not file or file.filename == ''):
        logger.error("Atendido es 'true' pero no se recibió archivo PDF.")
        return jsonify({"error": "Debe adjuntar un PDF si marca Atendido"}), 400

    connection = get_db_connection()
    if connection is None:
        logger.error("Error al conectar con la base de datos.")
        return jsonify({"error": "Error al conectar a la base de datos"}), 500

    try:
        cursor = connection.cursor(dictionary=True)
        logger.info(f"Consultando registro con id={row_id}")
        cursor.execute("""
            SELECT dp_plazos.id,
                   dp_plazos.registro_ppu,
                   dp.abogado,
                   dp.origen
            FROM datapenal_plazos dp_plazos
            JOIN datapenal dp ON dp_plazos.registro_ppu = dp.registro_ppu
            WHERE dp_plazos.id = %s
            LIMIT 1
        """, (row_id,))
        row = cursor.fetchone()
        if not row:
            logger.error(f"No se encontró registro en datapenal_plazos con id={row_id}")
            return jsonify({"error": f"No se encontró registro en datapenal_plazos con id={row_id}"}), 404

        registro_ppu = row.get('registro_ppu') or 'SIN_PPU'
        abogado = row.get('abogado') or 'SIN_ABOGADO'
        origen = row.get('origen') or 'SIN_ORIGEN'
        logger.info(f"Registro encontrado: registro_ppu={registro_ppu}, abogado={abogado}, origen={origen}")

        # Solo se actualiza si se envía un PDF válido
        if atendido_str == 'true' and file and file.filename != '':
            logger.info("Procesando archivo PDF adjunto.")
            months_map = {
                1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL",
                5: "MAYO", 6: "JUNIO", 7: "JULIO", 8: "AGOSTO",
                9: "SETIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"
            }
            today = datetime.now()
            year_str = str(today.year)
            month_str = months_map.get(today.month, "MES_DESCONOCIDO")
            day_str = today.strftime("%d-%m-%Y")
            final_folder = os.path.join(DESTINATION_BASE_PATH, year_str, month_str, day_str)
            logger.debug(f"Carpeta destino: {final_folder}")
            os.makedirs(final_folder, exist_ok=True)
            logger.info("Carpeta destino verificada/creada.")

            # Construir el nuevo nombre conservando los espacios
            original_filename = file.filename  # Se conservan los espacios
            ext = os.path.splitext(original_filename)[1].lower()
            new_filename = f"ESCRITO {abogado.strip()} {registro_ppu.strip()} {origen.strip() if origen else 'NOCASO'}{ext}"
            final_path = os.path.join(final_folder, new_filename)
            logger.info(f"Nuevo nombre de archivo: {new_filename}")
            logger.debug(f"Ruta final calculada: {final_path}")

            try:
                file.save(final_path)
                logger.info(f"Archivo guardado en: {final_path}")
            except Exception as save_err:
                logger.error(f"Error al guardar el archivo: {save_err}", exc_info=True)
                connection.rollback()
                return jsonify({"error": "Error al guardar el PDF en el destino final."}), 500

            # Verificar de forma explícita que el archivo se encuentre en la ruta destino
            if not os.path.isfile(final_path):
                logger.error(f"Verificación fallida: el archivo no se encontró en {final_path}")
                connection.rollback()
                return jsonify({"error": "Error: el PDF no se movió al destino final."}), 400

            # Calcular el hash SHA-256 del archivo guardado
            try:
                sha256 = hashlib.sha256()
                with open(final_path, 'rb') as f:
                    for chunk in iter(lambda: f.read(4096), b""):
                        sha256.update(chunk)
                hash_respuesta = sha256.hexdigest()
                logger.info(f"Hash SHA-256 calculado: {hash_respuesta}")
            except Exception as hash_err:
                logger.error(f"Error al calcular hash: {hash_err}", exc_info=True)
                connection.rollback()
                return jsonify({"error": "Error al calcular el hash del PDF."}), 500

            # Actualizar la BD: únicamente si el archivo se movió y existe, se actualiza seguimiento a 'ATENDIDA'
            update_query_pdf = """
                UPDATE datapenal_plazos
                SET seguimiento = %s,
                    ruta_respuesta = %s,
                    hash_respuesta = %s,
                    last_modified = NOW()
                WHERE id = %s
            """
            cursor.execute(update_query_pdf, ("ATENDIDA", final_path, hash_respuesta, row_id))
            logger.info("Registro actualizado con seguimiento = 'ATENDIDA', ruta_respuesta y hash_respuesta.")
            ruta_respuesta = final_path
        else:
            logger.error("Condición de PDF no cumplida: se esperaba un archivo PDF.")
            connection.rollback()
            return jsonify({"error": "No se proporcionó el PDF necesario para actualizar seguimiento."}), 400

        connection.commit()
        logger.info("Transacción comprometida correctamente.")
        return jsonify({
            "message": "Seguimiento marcado como ATENDIDA y PDF subido correctamente.",
            "ruta_respuesta": ruta_respuesta,
            "hash_respuesta": hash_respuesta
        }), 200

    except Exception as e:
        connection.rollback()
        logger.error(f"Excepción en actualizar_seguimiento: {e}", exc_info=True)
        return jsonify({"error": "Error al actualizar seguimiento"}), 500

    finally:
        cursor.close()
        connection.close()



        
@app.route('/api/buscar_rango', methods=['GET'])
@login_required
def buscar_rango():
    """
    Busca registros dentro de un rango [ppu_inicio, ppu_fin].
    """
    ppu_inicio = request.args.get('ppu_inicio', '').strip()
    ppu_fin = request.args.get('ppu_fin', '').strip()
    query = request.args.get('query', '').strip()

    # ===========================
    # FORZAR ABOGADO para "user"
    user_role = session.get('role')
    current_username = session.get('username', '')
    if user_role == 'user':
        abogado_filter = username_to_abogado.get(current_username, '').upper()
    else:
        abogado_filter = request.args.get('abogado', '').strip().upper()
    # ===========================

    mostrar_archivados = request.args.get('mostrar_archivados', 'true').lower() == 'true'

    connection = get_db_connection()
    if connection is None:
        return jsonify({"error": "Error al conectar con la base de datos"}), 500

    try:
        cursor = connection.cursor(dictionary=True)
        conditions = []
        params = []

        if not mostrar_archivados:
            conditions.append("(etiqueta IS NULL OR etiqueta != %s)")
            params.append('ARCHIVO')

        if abogado_filter:
            selected_abogado_regex = f"(?i)\\b{re.escape(abogado_filter)}\\b"
            conditions.append("abogado REGEXP %s")
            params.append(selected_abogado_regex)

        if query:
            columns = [
                'abogado','denunciado','origen','delito','departamento',
                'fiscalia','informe_juridico','item','e_situacional','registro_ppu','juzgado','etiqueta'
            ]
            regexp_query = query_to_regexp(query)
            regexp_conditions = []
            regexp_params = []
            for col in columns:
                regexp_conditions.append(f"{col} REGEXP %s")
                regexp_params.append(regexp_query)
            if regexp_conditions:
                conditions.append('(' + ' OR '.join(regexp_conditions) + ')')
                params.extend(regexp_params)

        where_clause = 'WHERE ' + ' AND '.join(conditions) if conditions else ''
        sql_query = f"SELECT * FROM datapenal {where_clause}"
        cursor.execute(sql_query, params)
        all_data = cursor.fetchall()

        # Ordenar en Python
        all_data_sorted = sorted(all_data, key=lambda r: parse_ppu(r['registro_ppu']))

        # Si no se especifica ni ppu_inicio ni ppu_fin, devuelves todo
        if not ppu_inicio and not ppu_fin:
            return jsonify({"data": all_data_sorted}), 200

        # Determinar tuplas de rango
        if ppu_inicio:
            start_tuple = parse_ppu(ppu_inicio)
        else:
            start_tuple = (0, 0, 0, '')
        if ppu_fin:
            end_tuple = parse_ppu(ppu_fin)
        else:
            end_tuple = (999999, 999999, 999999, 'ZZZZ')

        year_min = start_tuple[1]
        year_max = end_tuple[1]

        filtered_data = []
        for row in all_data_sorted:
            row_tuple = parse_ppu(row['registro_ppu'])
            row_year = row_tuple[1]

            if row_year < year_min or row_year > year_max:
                continue
            if row_tuple >= start_tuple and row_tuple <= end_tuple:
                filtered_data.append(row)

        return jsonify({"data": filtered_data}), 200

    except Exception as e:
        print(f"Error en buscar_rango: {e}")
        return jsonify({"error": "Error interno en buscar_rango"}), 500
    finally:
        if cursor:
            cursor.close()
        connection.close()

@app.route('/api/notifications', methods=['GET'])
@login_required
def get_notifications():
    if session.get('role') != 'user':
        return jsonify({"error": "No autorizado"}), 403

    connection = get_db_connection()
    if connection is None:
        return jsonify({"error": "Error al conectar con la base de datos"}), 500

    try:
        cursor = connection.cursor(dictionary=True)
        current_username = session.get('username', '')
        abogado = username_to_abogado.get(current_username, '').upper()

        query = """
        SELECT 
            registro_ppu,
            SUM(CASE WHEN leido = 0 THEN 1 ELSE 0 END) AS unread_count,
            MIN(fecha_version) AS fecha_version_min
        FROM datapenal_versioning
        WHERE abogado = %s AND fecha_version >= '2025-02-02'
        GROUP BY registro_ppu
        ORDER BY fecha_version_min ASC
        """
        cursor.execute(query, (abogado,))
        notifications = cursor.fetchall()

        if not notifications:
            return jsonify({"message": "No hay notificaciones pendientes."}), 200

        return jsonify({"notifications": notifications}), 200
    except Exception as e:
        return jsonify({"error": f"Error al obtener notificaciones: {e}"}), 500
    finally:
        cursor.close()
        connection.close()


@app.route('/api/notifications/mark_read', methods=['POST'])
@login_required
def mark_notifications_read():
    if session.get('role') != 'user':
        return jsonify({"error": "No autorizado"}), 403

    data = request.json
    registro_ppu = data.get("registro_ppu")
    leido = data.get("leido")  # Determinar si marcar como leido (1) o no leido (0)
    if not registro_ppu or leido is None:
        return jsonify({"error": "El registro_ppu y el estado de lectura son requeridos"}), 400

    connection = get_db_connection()
    if connection is None:
        return jsonify({"error": "Error al conectar con la base de datos"}), 500

    try:
        cursor = connection.cursor()
        cursor.execute(
            "UPDATE datapenal_versioning SET leido = %s WHERE registro_ppu = %s", 
            (leido, registro_ppu)
        )
        connection.commit()
        return jsonify({"message": "Notificaciones actualizadas"}), 200
    except Exception as e:
        return jsonify({"error": f"Error al actualizar notificaciones: {e}"}), 500
    finally:
        cursor.close()
        connection.close()

@app.route('/api/notifications/mark_read_individual', methods=['POST'])
@login_required
def mark_notification_read_individual():
    if session.get('role') != 'user':
        return jsonify({"error": "No autorizado"}), 403

    data = request.json
    id = data.get("id")  # Obtiene el id de la notificación
    leido = data.get("leido")  # Obtiene el estado de lectura (1 o 0)

    # Validar si id y leido están presentes
    if not id or leido is None:
        return jsonify({"error": "El id y el estado de lectura son requeridos"}), 400

    connection = get_db_connection()
    if connection is None:
        return jsonify({"error": "Error al conectar con la base de datos"}), 500

    try:
        cursor = connection.cursor()
        cursor.execute(
            "UPDATE datapenal_versioning SET leido = %s WHERE id = %s",  # Se utiliza id
            (leido, id)
        )
        connection.commit()
        return jsonify({"message": "Notificación actualizada"}), 200
    except Exception as e:
        return jsonify({"error": f"Error al actualizar notificación: {e}"}), 500
    finally:
        cursor.close()
        connection.close()





@app.route('/api/notifications/details', methods=['GET'])
@login_required
def notification_details():
    if session.get('role') != 'user':
        return jsonify({"error": "No autorizado"}), 403

    registro_ppu = request.args.get("registro_ppu", "").strip()
    if not registro_ppu:
        return jsonify({"error": "El registro_ppu es requerido"}), 400

    connection = get_db_connection()
    if connection is None:
        return jsonify({"error": "Error al conectar con la base de datos"}), 500

    try:
        cursor = connection.cursor(dictionary=True)
        current_username = session.get('username', '')
        abogado = username_to_abogado.get(current_username, '').upper()

        cursor.execute("""
            SELECT 
                id, version_id, abogado, registro_ppu, denunciado, origen, juzgado, fiscalia, departamento,
                e_situacional, DATE_FORMAT(fecha_version, '%d-%m-%Y') AS fecha_version, 
                usuario_modificacion, ruta, leido
            FROM datapenal_versioning
            WHERE registro_ppu = %s AND abogado = %s
            ORDER BY fecha_version DESC
        """, (registro_ppu, abogado))
        details = cursor.fetchall()
        return jsonify({"details": details}), 200
    except Exception as e:
        return jsonify({"error": f"Error al obtener detalles: {e}"}), 500
    finally:
        cursor.close()
        connection.close()

@app.route('/api/descargar_pdf_minimal', methods=['GET'])
@login_required
def descargar_pdf_minimal():
    # Se intenta obtener el parámetro 'filename'
    filename = request.args.get('filename')
    # Si 'filename' no es válido o es "undefined", se intenta obtener 'ruta'
    if not filename or filename.lower() == "undefined":
        filename = request.args.get('ruta')
    # Si aún no se obtuvo un valor válido, se retorna error
    if not filename or filename.lower() == "undefined":
        return jsonify({"error": "Se requiere un nombre de archivo válido"}), 400

    temp_folder = os.path.abspath(app.config['UPLOAD_FOLDER'])
    file_path = os.path.join(temp_folder, filename)

    if not os.path.isfile(file_path):
        return jsonify({"error": "Archivo no encontrado"}), 404

    try:
        response = send_file(file_path, as_attachment=False)
        response.headers["Content-Disposition"] = f'inline; filename="{filename}"'
        response.headers["X-Window-Target"] = "pdfWindow"
        return response
    except Exception as e:
        return jsonify({"error": f"Error al obtener el archivo: {str(e)}"}), 500

    # Nuevo endpoint para predecir mediante Machine Learning


@app.route('/api/check_duplicate_situacion', methods=['POST'])
@login_required
def check_duplicate_situacion():
    """
    Recibe un JSON con 'registro_ppu' y 'e_situacional' y consulta en la tabla datapenal_versioning
    para verificar si ya existe un registro para ese 'registro_ppu' cuyo 'e_situacional' coincida (ignorando mayúsculas)
    y filtrando para que 'ruta' y 'hash_sha' no sean NULL.
    Ejemplo de body:
    {
      "registro_ppu": "D-123-2023",
      "e_situacional": "Disposición N° 03"
    }
    """
    data = request.json
    registro_ppu = data.get('registro_ppu', '').strip()
    e_situacional = data.get('e_situacional', '').strip()

    if not registro_ppu or not e_situacional:
        return jsonify({"error": "Se requieren los campos 'registro_ppu' y 'e_situacional'"}), 400

    connection = get_db_connection()
    if not connection:
        return jsonify({"error": "Error al conectar con la base de datos"}), 500

    try:
        # Construir búsqueda: se fuerza que se busque para ese registro_ppu
        text_search = f"%{e_situacional.lower()}%"
        cursor = connection.cursor(dictionary=True)
        query = """
            SELECT COUNT(*) AS conteo
            FROM datapenal_versioning
            WHERE registro_ppu = %s
              AND ruta IS NOT NULL
              AND hash_sha IS NOT NULL
              AND LOWER(e_situacional) LIKE %s
        """
        cursor.execute(query, (registro_ppu, text_search))
        row = cursor.fetchone()
        duplicate_count = row['conteo']
        if duplicate_count > 0:
            return jsonify({
                "is_duplicate": True,
                "count": duplicate_count,
                "message": "Se encontraron coincidencias en la BD para el registro_ppu dado."
            }), 200
        else:
            return jsonify({
                "is_duplicate": False,
                "count": 0,
                "message": "No se encontraron coincidencias."
            }), 200
    except Exception as e:
        return jsonify({"error": f"Error al buscar duplicados: {e}"}), 500
    finally:
        cursor.close()
        connection.close()

@app.route('/api/historiales', methods=['POST'])
@login_required
def obtener_historiales():
    data = request.json
    ppus = data.get('registro_ppu', [])
    if not ppus:
        return jsonify({"error": "No se proporcionaron registros PPU"}), 400

    connection = get_db_connection()
    if connection is None:
        return jsonify({"error": "Error al conectar con la base de datos"}), 500

    try:
        cursor = connection.cursor(dictionary=True)
        format_strings = ','.join(['%s'] * len(ppus))
        sql = f"""
            SELECT registro_ppu, version_id, abogado, denunciado, origen, juzgado, 
                   fiscalia, departamento, e_situacional, DATE_FORMAT(fecha_version, '%d-%m-%Y') AS fecha_version, 
                   usuario_modificacion, ruta
            FROM datapenal_versioning
            WHERE registro_ppu IN ({format_strings})
            ORDER BY registro_ppu, fecha_version DESC
        """
        cursor.execute(sql, tuple(ppus))
        rows = cursor.fetchall()

        # Filtrar solo aquellos registros donde 'ruta' tenga un valor válido:
        rows_filtrados = [
            row for row in rows 
            if row.get('ruta') and row['ruta'].strip() and row['ruta'].strip().upper() != "NULL"
        ]

        # Organizar los historiales en un diccionario
        historiales = {}
        for row in rows_filtrados:
            ppu = row['registro_ppu']
            if ppu not in historiales:
                historiales[ppu] = []
            historiales[ppu].append(row)

        return jsonify({"historiales": historiales}), 200
    except Exception as e:
        return jsonify({"error": f"Error al obtener historiales: {str(e)}"}), 500
    finally:
        cursor.close()
        connection.close()






@app.route('/logs', methods=['GET'])
@login_required
def view_logs():
    """
    Devuelve el contenido del archivo de log en formato HTML, con la capacidad
    de actualizar en tiempo real (aproximadamente cada 1 segundo).
    """
    try:
        # Leer el archivo de log y obtener la última línea
        with open(log_path, 'r', encoding='utf-8', errors='replace') as f:
            log_content = f.read()

        # HTML para visualizarlo de manera adecuada
        html_content = f"<html><body><pre>{log_content}</pre></body></html>"

        # Devolver el contenido actualizado
        return html_content
    except Exception as e:
        return jsonify({"error": f"No se pudo leer el archivo de log: {str(e)}"}), 500


@app.route('/api/eliminar_fila', methods=['POST'])
@login_required
@role_required(['admin'])
def eliminar_fila():
    data = request.json
    row_id = data.get('id')
    if not row_id:
        logger.error("El parámetro 'id' es requerido para eliminar la fila.")
        return jsonify({"error": "ID es requerido para eliminar la fila."}), 400

    connection = get_db_connection()
    if connection is None:
        logger.error("Error al conectar con la base de datos.")
        return jsonify({"error": "Error al conectar a la base de datos."}), 500

    try:
        cursor = connection.cursor()
        # Primero, eliminar las filas de alertas_enviadas que dependan del registro a borrar
        cursor.execute("DELETE FROM alertas_enviadas WHERE plazos_id = %s", (row_id,))
        logger.info(f"Dependencias en 'alertas_enviadas' eliminadas para plazos_id = {row_id}.")

        # Luego, eliminar la fila en datapenal_plazos
        cursor.execute("DELETE FROM datapenal_plazos WHERE id = %s", (row_id,))
        connection.commit()
        logger.info(f"Fila con id {row_id} eliminada exitosamente.")
        return jsonify({"message": "Fila eliminada exitosamente."}), 200
    except Exception as e:
        connection.rollback()
        logger.error(f"Error al eliminar la fila con id {row_id}: {e}", exc_info=True)
        return jsonify({"error": "Error al eliminar la fila."}), 500
    finally:
        cursor.close()
        connection.close()

################################################################################################################################################
### CONSULTAS PENALES###########################################################################################################################
################################################################################################################################################


@app.route('/api/consulta_ppupenal', methods=['GET'])
@login_required
def get_consulta_ppupenal():
    connection = get_db_connection()
    if connection is None:
        return jsonify({"success": False, "message": "Error al conectar con la base de datos"}), 500
    try:
        cursor = connection.cursor(dictionary=True)
        query = """
            SELECT 
                consulta_ppu,
                abogado,
                denunciado,
                origen,
                `nr de exp completo`,
                fiscalia,
                departamento,
                juzgado,
                delito,
                informe_juridico,
                item,
                e_situacional,
                fecha_ingreso,
                fecha_e_situacional,
                etiqueta,
                registro_ppu,
                decision_coordinador
            FROM consulta_ppupenal
        """
        cursor.execute(query)
        rows = cursor.fetchall()
        return jsonify({"success": True, "constulas": rows}), 200
    except Exception as e:
        return jsonify({"success": False, "message": f"Error al obtener datos: {str(e)}"}), 500
    finally:
        cursor.close()
        connection.close()


@app.route('/api/generar_registro_consulta', methods=['POST'])
@login_required
@role_required(['admin'])
def generar_consulta():
    try:
        data = request.json
        logger.info("Datos recibidos en la solicitud: %s", data)

        year = data.get('year')
        caso_especial = data.get('caso_especial', False)

        if not year:
            logger.error("Error: Año no proporcionado")
            return jsonify({"error": "Año es requerido"}), 400

        if caso_especial:
            numero = data.get('numero')
            sufijo = data.get('sufijo', '')
            logger.info("Procesando caso especial: número=%s, sufijo=%s", numero, sufijo)

            if not numero:
                logger.error("Error: Número requerido en caso especial")
                return jsonify({"error": "Número es requerido para casos especiales"}), 400

            registro_ppu = f"CONS-{numero}-{year}"
            if sufijo:
                registro_ppu += f"-{sufijo}"
            logger.info("Registro generado (especial): %s", registro_ppu)
            return jsonify({"registro_ppu": registro_ppu, "success": True})

        # Caso general
        connection = get_db_connection()
        if connection is None:
            logger.error("Error: Conexión a la base de datos fallida")
            return jsonify({"error": "Error al conectar con la base de datos"}), 500

        try:
            cursor = connection.cursor(dictionary=True)
            like_pattern = f"CONS-%-{year}"
            logger.info("Ejecutando consulta con patrón LIKE: %s", like_pattern)

            cursor.execute(
                "SELECT TRIM(consulta_ppu) AS consulta_ppu FROM consulta_ppupenal WHERE consulta_ppu LIKE %s",
                (like_pattern,)
            )

            registros = cursor.fetchall()
            logger.info("Registros obtenidos: %s", registros)

            numeros = []
            for reg in registros:
                consulta_ppu_val = reg['consulta_ppu'].strip()
                match = re.match(rf'^CONS-(\d+)-{year}$', consulta_ppu_val)
                if match:
                    numeros.append(int(match.group(1)))
            logger.info("Números extraídos: %s", numeros)

            next_num = max(numeros) + 1 if numeros else 1
            registro_ppu = f"CONS-{next_num:03d}-{year}"
            logger.info("Registro generado: %s", registro_ppu)

            return jsonify({"registro_ppu": registro_ppu, "success": True})

        except Exception as e:
            logger.error("Error al generar registro en bloque try: %s", e)
            return jsonify({"error": f"Error al generar registro: {e}"}), 500

        finally:
            cursor.close()
            connection.close()

    except Exception as e:
        logger.error("Error inesperado en endpoint: %s", e)
        return jsonify({"error": f"Error inesperado: {e}"}), 500



@app.route('/api/agregar_consulta', methods=['POST'])
@login_required
@role_required(['admin'])
def agregar_caso_consulta():
    data = request.json
    logger.info("Datos recibidos en agregar_consulta: %s", data)

    expediente_juzgado = data.pop('expediente_juzgado', None)
    nr_de_exp = data.get('nr de exp completo', '').strip()
    logger.debug("Número de expediente transformado: %s", nr_de_exp)

    if nr_de_exp:
        connection = get_db_connection()
        if connection is None:
            logger.error("Error al conectar con la base de datos (duplicados)")
            return jsonify({"error": "Error al conectar con la base de datos"}), 500
        try:
            cursor = connection.cursor()
            query_check = "SELECT COUNT(*) FROM datapenal WHERE `nr de exp completo` = %s"
            cursor.execute(query_check, (nr_de_exp,))
            result_datapenal = cursor.fetchone()[0]
            logger.debug("Duplicados en datapenal: %s", result_datapenal)
            query_check_consulta = "SELECT COUNT(*) FROM consulta_ppupenal WHERE `nr de exp completo` = %s"
            cursor.execute(query_check_consulta, (nr_de_exp,))
            result_consulta = cursor.fetchone()[0]
            logger.debug("Duplicados en consulta_ppupenal: %s", result_consulta)
            if result_datapenal > 0 or result_consulta > 0:
                logger.error("Número de expediente ya registrado: %s", nr_de_exp)
                return jsonify({"error": f"El número de expediente '{nr_de_exp}' ya está registrado."}), 400
        except Exception as e:
            logger.error("Error al verificar duplicados: %s", e, exc_info=True)
            return jsonify({"error": "Error al verificar duplicados"}), 500
        finally:
            cursor.close()
            connection.close()

    if expediente_juzgado:
        if not isinstance(expediente_juzgado, dict):
            logger.error("expediente_juzgado no es un objeto: %s", expediente_juzgado)
            return jsonify({"error": "El campo 'expediente_juzgado' debe ser un objeto."}), 400
        errores = validate_expediente_juzgado(expediente_juzgado)
        if errores:
            logger.error("Errores en expediente_juzgado: %s", errores)
            return jsonify({"error": errores}), 400
        expediente_formateado = (
            f"Exp. {expediente_juzgado['campo1']}-"
            f"{expediente_juzgado['campo2']}-"
            f"{expediente_juzgado['campo3']}-"
            f"{expediente_juzgado['campo4']}-"
            f"{expediente_juzgado['campo5']}-"
            f"{expediente_juzgado['campo6']}-"
            f"{expediente_juzgado['campo7']}"
        )
        origen = f"{data.get('origen', '')}, {expediente_formateado}" if data.get('origen') else expediente_formateado
        data['origen'] = origen
    else:
        if 'origen' in data and data['origen']:
            if data['origen'][0].isdigit():
                data['origen'] = 'CASO ' + data['origen']

    data['fecha_ingreso'] = datetime.now().date()
    logger.debug("Datos a insertar en consulta_ppupenal: %s", data)
    connection = get_db_connection()
    if connection is None:
        logger.error("Error al conectar con la base de datos (INSERT)")
        return jsonify({"error": "Error al conectar con la base de datos"}), 500
    try:
        cursor = connection.cursor()
        columns = ', '.join(f"`{k}`" for k in data.keys())
        placeholders = ', '.join('%s' for _ in data.values())
        values = tuple(data.values())
        query = f'INSERT INTO consulta_ppupenal ({columns}) VALUES ({placeholders})'
        logger.info("Ejecutando query: %s con valores %s", query, values)
        cursor.execute(query, values)
        connection.commit()
        logger.info("Caso agregado exitosamente a consulta_ppupenal")
    except Exception as e:
        logger.error("Error al agregar caso: %s", e, exc_info=True)
        return jsonify({"error": "Error al agregar caso"}), 500
    finally:
        cursor.close()
        connection.close()

    # Lógica de versioning para consulta
    original_file_name = (data.get('fileName') or data.get('ruta') or '').strip()
    if not original_file_name:
        logger.warning("Nombre de archivo no proporcionado; se omite versioning.")
        return jsonify({"message": "Caso agregado sin versioning por falta de nombre de archivo"}), 200

    pdf_path = os.path.join(app.config['UPLOAD_FOLDER'], original_file_name)
    if not os.path.isfile(pdf_path):
        logger.error("Archivo PDF no encontrado: %s", pdf_path)
        return jsonify({"error": f"Archivo PDF no encontrado: {pdf_path}"}), 400

    months_map = {
        1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL",
        5: "MAYO", 6: "JUNIO", 7: "JULIO", 8: "AGOSTO",
        9: "SETIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"
    }
    today = datetime.now()
    year_str = str(today.year)
    month_str = months_map.get(today.month, "MES_DESCONOCIDO")
    day_str = today.strftime("%d-%m-%Y")
    final_folder = os.path.join(DESTINATION_BASE_PATH, year_str, month_str, day_str)
    os.makedirs(final_folder, exist_ok=True)

    dest_path = os.path.join(final_folder, original_file_name)
    try:
        with open(pdf_path, 'rb') as src, open(dest_path, 'wb') as dst:
            shutil.copyfileobj(src, dst, length=1048576)
        if not os.path.isfile(dest_path):
            logger.error("Error al copiar el archivo a %s", dest_path)
            return jsonify({"error": f"Error al copiar el archivo a {dest_path}"}), 500
        logger.info("PDF copiado a: %s", dest_path)
    except Exception as e:
        logger.error("Error al copiar %s a %s: %s", pdf_path, dest_path, e)
        return jsonify({"error": f"Error al copiar el archivo: {e}"}), 500

    def sha256_of_file(filepath):
        sha256 = hashlib.sha256()
        try:
            with open(filepath, 'rb') as f:
                for chunk in iter(lambda: f.read(4096), b""):
                    sha256.update(chunk)
            return sha256.hexdigest()
        except Exception as e:
            logger.error("Error al calcular SHA-256 del archivo %s: %s", filepath, e)
            return None

    hash_sha = sha256_of_file(dest_path)
    if not hash_sha:
        logger.error("No se pudo calcular el hash para %s", dest_path)
        return jsonify({"error": "Error al calcular hash del PDF"}), 500

    versioning_data = {
        'abogado': data.get('abogado', ''),
        'consulta_ppu': data.get('consulta_ppu', ''),
        'denunciado': data.get('denunciado', ''),
        'origen': data.get('origen', ''),
        'fiscalia': data.get('fiscalia', ''),
        'juzgado': data.get('juzgado', ''),
        'departamento': data.get('departamento', ''),
        'e_situacional': data.get('e_situacional', ''),
        'fecha_version': datetime.now(),
        'usuario_modificacion': session.get('username', 'desconocido'),
        'ruta': os.path.normpath(dest_path),
        'hash_sha': hash_sha
    }
    connection = get_db_connection()
    if connection is None:
        logger.error("Error al conectar con la base de datos para versioning")
        return jsonify({"error": "Error al conectar con la base de datos para versioning"}), 500
    try:
        cursor = connection.cursor()
        cols_ver = ', '.join(f"`{k}`" for k in versioning_data.keys())
        placeholders_ver = ', '.join('%s' for _ in versioning_data.values())
        values_ver = tuple(versioning_data.values())
        query_ver = f"INSERT INTO datapenal_versioning ({cols_ver}) VALUES ({placeholders_ver})"
        cursor.execute(query_ver, values_ver)
        connection.commit()
        logger.info("Registro agregado en datapenal_versioning para consulta")
    except Exception as e:
        logger.error("Error al insertar en datapenal_versioning: %s", e, exc_info=True)
        return jsonify({"error": "Error al insertar en datapenal_versioning"}), 500
    finally:
        cursor.close()
        connection.close()

    return jsonify({"message": "Caso agregado y versioning registrado"}), 200



@app.route('/api/get_registros_consulta', methods=['GET'])
@login_required
def get_registros_consulta():
    year = request.args.get('year')
    if not year:
        return jsonify({"error": "Año es requerido"}), 400
    connection = get_db_connection()
    if connection is None:
        return jsonify({"error": "Error al conectar con la base de datos"}), 500
    try:
        cursor = connection.cursor(dictionary=True)
        regex_pattern = r'^CONS-0*\d+-{}(?:-[A-Z]+)?$'.format(year)
        cursor.execute(
            "SELECT consulta_ppu FROM consulta_ppupenal WHERE consulta_ppu REGEXP %s ORDER BY consulta_ppu ASC",
            (regex_pattern,)
        )
        registros = cursor.fetchall()
        registros_list = [reg['consulta_ppu'] for reg in registros]
        return jsonify({"data": registros_list})
    except Exception as e:
        print(f"Error al obtener registros: {e}")
        return jsonify({"error": f"Error al obtener registros: {e}"}), 500
    finally:
        cursor.close()
        connection.close()


@app.route('/api/eliminar_consulta', methods=['POST'])
@login_required
@role_required(['admin'])
def eliminar_consulta():
    data = request.json
    consulta_ppu = data.get('consulta_ppu')

    if not consulta_ppu:
        return jsonify({"error": "Consulta PPU es requerida"}), 400

    connection = get_db_connection()
    if connection is None:
        return jsonify({"error": "Error al conectar con la base de datos"}), 500

    try:
        cursor = connection.cursor()
        cursor.execute("DELETE FROM datappupenal.consulta_ppupenal WHERE consulta_ppu = %s", (consulta_ppu,))
        connection.commit()
        return jsonify({"message": "Caso eliminado exitosamente"}), 200
    except Exception as e:
        print(f"Error al eliminar caso: {e}")
        return jsonify({"error": "Error al eliminar caso"}), 500
    finally:
        cursor.close()
        connection.close()


def get_table_columns(connection, table_name):
    """
    Retorna un conjunto con los nombres de las columnas de la tabla indicada.
    """
    cur = connection.cursor()
    cur.execute(f"SHOW COLUMNS FROM {table_name}")
    cols = {row[0] for row in cur.fetchall()}
    cur.close()
    return cols

@app.route('/api/decision_coordinador', methods=['POST'])
def decision_coordinador():
    """
    Endpoint que ejecuta las operaciones:
      • ACUMULACION  
      • PALANCA  
      • INGRESO NUEVO

    Se recibe en el body JSON:
      - consulta_ppu (obligatorio; de consulta_ppupenal)
      - operation: "ACUMULACION", "PALANCA" o "INGRESO_NUEVO" (obligatorio)
      - Para ACUMULACION: acumulacionValue  
      - Para PALANCA: palancaLegajo  
      - Para INGRESO NUEVO: ingresoNuevoOption ("DENUNCIA" o "LEGAJO") y anio
    """
    data = request.json
    consulta_ppu = data.get("consulta_ppu", "").strip()
    operation = data.get("operation", "").strip().upper()

    if not consulta_ppu or not operation:
        return jsonify({"success": False, "message": "Parámetros 'consulta_ppu' y 'operation' son requeridos."}), 400

    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({"success": False, "message": "Error al conectar con la base de datos."}), 500

        # Se utiliza un cursor con resultado en diccionario para facilitar el manejo de los campos.
        cursor = connection.cursor(dictionary=True)

        if operation == "ACUMULACION":
            acumulacion_value = data.get("acumulacionValue", "").strip()
            if not acumulacion_value:
                return jsonify({"success": False, "message": "El valor para Acumulación es requerido."}), 400

            update_queries = [
                ("UPDATE datapenal SET registro_ppu = %s WHERE registro_ppu LIKE %s", 
                 (acumulacion_value, f"%{consulta_ppu}%")),
                ("UPDATE consulta_ppupenal SET registro_ppu = %s WHERE consulta_ppu = %s", 
                 (acumulacion_value, consulta_ppu)),
                ("UPDATE datapenal_versioning SET registro_ppu = %s WHERE consulta_ppu = %s", 
                 (acumulacion_value, consulta_ppu)),
                ("UPDATE datapenal_plazos SET registro_ppu = %s WHERE consulta_ppu = %s", 
                 (acumulacion_value, consulta_ppu))
            ]
            for q, params in update_queries:
                cursor.execute(q, params)
            connection.commit()
            return jsonify({"success": True, "message": "Operación Acumulación completada exitosamente."}), 200

        elif operation == "PALANCA":
            palanca_legajo = data.get("palancaLegajo", "").strip()
            if not palanca_legajo:
                return jsonify({"success": False, "message": "El número de legajo es requerido para Palanca."}), 400

            current_year = datetime.now().year
            formatted_legajo = format_legajo(palanca_legajo)
            palanca_key = f"PALANCA-{formatted_legajo}-{current_year}"
            # Se actualiza únicamente el campo palanca_ppu y la fecha de palanca, sin modificar registro_ppu.
            cursor.execute(
                "UPDATE consulta_ppupenal SET palanca_ppu = %s, fecha_palanca = NOW() WHERE consulta_ppu = %s",
                (palanca_key, consulta_ppu)
            )
            connection.commit()
            return jsonify({"success": True, "message": "Operación Palanca completada exitosamente.", "palanca": palanca_key}), 200


        elif operation == "INGRESO_NUEVO":
            ingreso_option = data.get("ingresoNuevoOption", "").strip().upper()
            anio_str = data.get("anio", "").strip()
            if not ingreso_option or not anio_str:
                return jsonify({"success": False, "message": "Se requieren 'ingresoNuevoOption' y 'anio' para Ingreso Nuevo."}), 400

            anio = int(anio_str)
            # Se define el prefijo y la expresión regular según la opción y el año.
            if ingreso_option == "DENUNCIA":
                prefix = "D-"
                regex_pattern = r'^D-0*\d+-{}(?:-[A-Z]+)?$'.format(anio)
            elif ingreso_option == "LEGAJO":
                if anio >= 2023:
                    prefix = "L. "
                    regex_pattern = r'^{}0*\d+-{}(?:-[A-Z]+)?$'.format(re.escape(prefix), anio)
                else:
                    prefix = "LEG-"
                    regex_pattern = r'^{}0*\d+-{}(?:-[A-Z]+)?$'.format(re.escape(prefix), anio)
            else:
                return jsonify({"success": False, "message": "Opción de Ingreso Nuevo no válida."}), 400

            # Se consulta la tabla datapenal para obtener los registros existentes que coincidan con el patrón.
            cursor.execute("SELECT registro_ppu FROM datapenal WHERE registro_ppu REGEXP %s", (regex_pattern,))
            registros = cursor.fetchall()
            numeros = []
            for reg in registros:
                reg_val = reg['registro_ppu']
                if ingreso_option == "LEGAJO":
                    patron_num = r'^{}0*(\d+)-{}'.format(re.escape(prefix), anio)
                else:
                    patron_num = r'^D-0*(\d+)-{}'.format(anio)
                match = re.match(patron_num, reg_val)
                if match:
                    numeros.append(int(match.group(1)))

            if ingreso_option == "LEGAJO" and anio >= 2023:
                if numeros:
                    numeros.sort()
                    for i in range(1, len(numeros)):
                        if numeros[i] > numeros[i - 1] + 1:
                            next_num = numeros[i - 1] + 1
                            break
                    else:
                        next_num = max(numeros) + 1
                else:
                    next_num = 1
            else:
                next_num = max(numeros) + 1 if numeros else 1

            max_length = len(str(next_num))
            new_registro = '{}{num:0{width}d}-{anio}'.format(prefix, num=next_num, width=max_length, anio=anio)

            # Se actualiza el registro_ppu en las tablas correspondientes, incluida consulta_ppupenal.
            update_queries = [
                ("UPDATE consulta_ppupenal SET registro_ppu = %s WHERE consulta_ppu = %s", (new_registro, consulta_ppu)),
                ("UPDATE datapenal_versioning SET registro_ppu = %s WHERE consulta_ppu = %s", (new_registro, consulta_ppu)),
                ("UPDATE datapenal_plazos SET registro_ppu = %s WHERE consulta_ppu = %s", (new_registro, consulta_ppu))
            ]
            for q, params in update_queries:
                cursor.execute(q, params)

            # Se obtiene la fila base de consulta_ppupenal para replicar la información en datapenal.
            cursor.execute("SELECT * FROM consulta_ppupenal WHERE consulta_ppu = %s LIMIT 1", (consulta_ppu,))
            existing_row = cursor.fetchone()
            if not existing_row:
                return jsonify({"success": False, "message": "No se encontró registro base para Ingreso Nuevo."}), 404

            # Se preparan las columnas a insertar en la tabla datapenal, filtrando las que correspondan.
            columnas = [desc['name'] for desc in cursor.description]
            new_row = dict(existing_row)
            new_row["registro_ppu"] = new_registro
            datapenal_cols = get_table_columns(connection, "datapenal")
            new_row_filtered = {k: v for k, v in new_row.items() if k in datapenal_cols}
            col_names = ", ".join(f"`{col}`" for col in new_row_filtered.keys() if col.lower() != "id")
            placeholders = ", ".join(["%s"] * (len(new_row_filtered) - (1 if any(col.lower() == "id" for col in new_row_filtered.keys()) else 0)))
            values = tuple(new_row_filtered[col] for col in new_row_filtered.keys() if col.lower() != "id")
            insert_query = f"INSERT INTO datapenal ({col_names}) VALUES ({placeholders})"
            cursor.execute(insert_query, values)

            connection.commit()
            return jsonify({"success": True, "message": f"Ingreso Nuevo ({ingreso_option}) completado.", "registro_ppu": new_registro}), 200

        else:
            return jsonify({"success": False, "message": "Operación no reconocida."}), 400

    except Exception as e:
        connection.rollback()
        app.logger.error("Error en /api/decision_coordinador: %s", e, exc_info=True)
        return jsonify({"success": False, "message": "Error interno del servidor."}), 500

    finally:
        cursor.close()
        connection.close()

## agregar logs de depuracion no genera bien el registro de legajos ni denuncias y ademas ver q actualice consulta_ppu y se pueda cammbiar abogado y otros daots en el fornt



################################################################################################################################################
### CONSULTAS PENALES###########################################################################################################################
################################################################################################################################################


################################################################################################################################################
### MODO BUSQUEDA TAB 0 ###########################################################################################################################
################################################################################################################################################




import re
from flask import jsonify, request
# Se asume que ya están definidos: get_db_connection, login_required, logger,
# parse_query_ppu, generar_variantes_ppu, normalizar_expediente, etc.

def build_legajo_regexp(query):
    """
    Construye un patrón REGEXP para buscar en el campo legajo según las siguientes reglas:
    
    - Se detecta si el query incluye un prefijo explícito ("D‑", "LEG‑", "L. " o "L.").
      Si es "D‑", la búsqueda se restringe a los legajos D‑;  
      si es "LEG‑" o "L." se permiten ambos formatos de forma intercambiable.
      
    - Se extrae el primer fragmento numérico (después del prefijo, hasta el primer guion).
      Se convierte a entero para ignorar ceros a la izquierda.
      
    - Si el query termina en guion (por ejemplo, "20-") se genera un patrón que exige
      que la parte numérica del legajo (después del prefijo) sea EXACTAMENTE la misma,
      es decir, "20-" forzará un match sólo si el legajo presenta como número 20 (aceptando
      variaciones como "20" o "020" siempre que al convertirlos a entero den 20), pero no "201".
      
    - Si el query no termina en guion (por ejemplo, "6" o "D-6"), se permite una comparación
      flexible (que busque coincidencias cuyo número inicie con el dígito(s) indicado).
      
    - Se incorpora opcionalmente el fragmento del año: si se especifica con 4 dígitos se fuerza
      la coincidencia exacta; si es menor a 4 dígitos se toma de forma "starts with".
      
    - Si el query es solamente el prefijo sin dígitos, se devuelve None para evitar búsquedas
      demasiado amplias.
    """
    q = query.strip().upper()
    # Detectar prefijo explícito (no se elimina si el usuario lo incluye, para restringir el grupo)
    prefix_match = re.match(r'^(D-|LEG-|L\. ?)', q)
    if prefix_match:
        raw_prefix = prefix_match.group(1).strip()
        if raw_prefix in ("LEG-", "L.", "L-"):
            explicit_prefix = "LEG"  # Tratar LEG y L. como equivalentes
        else:
            explicit_prefix = "D"
        remainder = q[len(prefix_match.group(1)):].strip()
    else:
        explicit_prefix = None
        remainder = q

    # Se exige que después (o en ausencia de prefijo) exista al menos un dígito;
    # de lo contrario se rechaza la búsqueda.
    if not re.match(r'^\d', remainder):
        return None

    # Si no aparece el guion, se asume que se quiere búsqueda flexible (se añade guion al final)
    if '-' not in remainder:
        remainder = remainder + "-"

    parts = remainder.split("-", 1)
    num_str = parts[0].strip()
    if not num_str.isdigit():
        return None
    num_val = int(num_str)
    # Flag exact: si el query termina en guion (sin contenido posterior al guion)
    exact = remainder.endswith("-")
    
    # Construir patrón para la parte numérica:
    if exact:
        # Se fuerza que no haya dígitos adicionales (lookahead negativo)
        num_pattern = rf'0*{num_val}(?!\d)'
    else:
        # Búsqueda flexible: el valor numérico debe empezar con el dígito(s)
        num_pattern = rf'0*{num_val}'
    
    # Procesar el fragmento de año, si lo hubiera
    year_pattern = ""
    if "-" in remainder:
        # Si hay contenido posterior al primer guion y no es vacío, se toma como año
        second_part = parts[1].strip()
        if second_part != "":
            if len(second_part) == 4 and second_part.isdigit():
                year_pattern = rf'-{second_part}\b'
            else:
                # Si la parte del año es menor a 4 dígitos, se usa de forma flexible (starts with)
                year_pattern = rf'-{second_part}'
        else:
            # Si es exactamente "num-" (sin datos del año), exigimos el guion solamente para marcar final del número
            year_pattern = "-"
    else:
        year_pattern = "-"
    
    # Construir el patrón para el prefijo:
    if explicit_prefix == "D":
        prefix_pattern = r'^D-'
    elif explicit_prefix == "LEG":
        prefix_pattern = r'^(?:LEG-|L\. ?)'
    else:
        prefix_pattern = r'^(?:D-|LEG-|L\. ?)'
        
    # Combinar todo
    pattern = prefix_pattern + num_pattern + year_pattern
    return pattern

@app.route("/api/new_search", methods=["GET"])
@login_required
def new_search():
    query = request.args.get("query", "").strip()
    search_field = request.args.get("search_field", "").strip()
    logger.info(f"new_search: 'query'='{query}', 'search_field'='{search_field}'")

    if not query:
        logger.warning("new_search: Parámetro 'query' vacío")
        return jsonify({"error": "El parámetro 'query' es obligatorio"}), 400

    if not search_field:
        logger.warning("new_search: Parámetro 'search_field' vacío")
        return jsonify({"error": "El parámetro 'search_field' es obligatorio"}), 400

    try:
        connection = get_db_connection()
        if connection is None:
            logger.error("new_search: Error al conectar con la base de datos")
            return jsonify({"error": "Error al conectar con la base de datos"}), 500

        cursor = connection.cursor(dictionary=True)

        # Se definen los patrones para búsquedas en campos relacionados a expedientes
        exp_pattern_1 = r'(\d{5}-\d{4}-\d{1,2}-\d{4}[A-Z]?-([A-Z]{2})-[A-Z]{2}-\d{1,2})'
        exp_pattern_2 = r'(\d{5}-\d{4}-\d{1,2}-[A-Z\d]+-[A-Z]{2}-[A-Z]{2}-\d{1,2})'

        # Mapeo de columnas en cada tabla
        field_map_datapenal = {
            "legajo": "registro_ppu",
            "casoFiscalCompleto": "`nr de exp completo`",
            "casoJudicial": "origen",
            "denunciado": "denunciado",
        }
        field_map_consulta = {
            "legajo": "consulta_ppu",
            "casoFiscalCompleto": "`nr de exp completo`",
            "casoJudicial": "origen",
            "denunciado": "denunciado",
        }

        if search_field not in field_map_datapenal:
            logger.warning(f"new_search: 'search_field' no soportado: {search_field}")
            return jsonify({"error": f"Campo de búsqueda no soportado: {search_field}"}), 400

        # Rama para el campo "legajo"
        if search_field == "legajo":
            pattern = build_legajo_regexp(query)
            if not pattern:
                logger.warning("new_search: Formato inválido en 'query' para 'legajo'")
                return jsonify({"error": "Formato de búsqueda inválido para registro PPU"}), 400

            sql_datapenal = f"""
                SELECT
                    *,
                    LENGTH({field_map_datapenal["legajo"]}) AS match_length,
                    'datapenal' AS source
                FROM datapenal
                WHERE {field_map_datapenal["legajo"]} REGEXP %s
                ORDER BY match_length ASC
                LIMIT 10
            """
            params_datapenal = (pattern,)
            sql_consulta = f"""
                SELECT
                    *,
                    LENGTH({field_map_consulta["legajo"]}) AS match_length,
                    'consulta' AS source
                FROM consulta_ppupenal
                WHERE {field_map_consulta["legajo"]} REGEXP %s
                ORDER BY match_length ASC
                LIMIT 10
            """
            params_consulta = (pattern,)
            cursor.execute(sql_datapenal, params_datapenal)
            results_datapenal = cursor.fetchall()
            cursor.execute(sql_consulta, params_consulta)
            results_consulta = cursor.fetchall()
            combined_results = results_datapenal + results_consulta

        # Rama para el campo "casoFiscalCompleto"
        elif search_field == "casoFiscalCompleto":
            query_sin_prefijo = query.upper().replace("CASO", "").strip()
            partes = query_sin_prefijo.split("-")
            if len(partes) >= 3 and partes[1].isdigit() and partes[2].isdigit():
                query_invertida = f"{partes[2]}-{partes[1]}"
            elif len(partes) == 2 and partes[0].isdigit() and partes[1].isdigit():
                query_invertida = f"{partes[1]}-{partes[0]}"
            else:
                query_invertida = query_sin_prefijo

            sql_datapenal = f"""
                SELECT
                    *,
                    LENGTH({field_map_datapenal["casoFiscalCompleto"]}) AS match_length,
                    'datapenal' AS source
                FROM datapenal
                WHERE {field_map_datapenal["casoFiscalCompleto"]} LIKE %s
                   OR {field_map_datapenal["casoFiscalCompleto"]} LIKE %s
                   OR origen LIKE %s
                   OR origen LIKE %s
            """
            params_datapenal = (
                f"%{query_sin_prefijo}%",
                f"%{query_invertida}%",
                f"%{query_sin_prefijo}%",
                f"%{query_invertida}%"
            )
            cursor.execute(sql_datapenal, params_datapenal)
            results_datapenal = cursor.fetchall()

            sql_consulta = f"""
                SELECT
                    *,
                    LENGTH({field_map_consulta["casoFiscalCompleto"]}) AS match_length,
                    'consulta' AS source
                FROM consulta_ppupenal
                WHERE {field_map_consulta["casoFiscalCompleto"]} LIKE %s
                   OR {field_map_consulta["casoFiscalCompleto"]} LIKE %s
                   OR origen LIKE %s
                   OR origen LIKE %s
            """
            params_consulta = (
                f"%{query_sin_prefijo}%",
                f"%{query_invertida}%",
                f"%{query_sin_prefijo}%",
                f"%{query_invertida}%"
            )
            cursor.execute(sql_consulta, params_consulta)
            results_consulta = cursor.fetchall()
            candidatos = results_datapenal + results_consulta

            patron_exp_flexible = r'\b\d{6,10}\s*-\s*\d{4}\s*-\s*\d{1,4}(?:\s*-\s*\d+)?\b'

            def partial_numeric_match_normal(candidate_value, query_value):
                try:
                    if "-" not in query_value:
                        m = re.fullmatch(r'\s*(\d+)\s*-\s*(\d+)\s*', candidate_value)
                        if not m:
                            return False
                        cand_first, cand_second = m.groups()
                        q_norm = str(int(query_value))
                        return cand_first.startswith(q_norm) or cand_second.startswith(q_norm)
                    else:
                        q_parts = query_value.split("-")
                        if len(q_parts) != 2:
                            return False
                        q_first = str(int(q_parts[0]))
                        q_second = q_parts[1].strip()
                        m = re.fullmatch(r'\s*(\d+)\s*-\s*(\d+)\s*', candidate_value)
                        if not m:
                            return False
                        cand_first, cand_second = m.groups()
                        return (cand_first == q_first) and (cand_second.startswith(q_second))
                except Exception:
                    return False

            def partial_numeric_match_inverted(candidate_value, query_value):
                try:
                    if "-" not in query_value:
                        m = re.fullmatch(r'\s*(\d+)\s*-\s*(\d+)\s*', candidate_value)
                        if not m:
                            return False
                        cand_first, cand_second = m.groups()
                        q_norm = str(int(query_value))
                        return cand_first.startswith(q_norm) or cand_second.startswith(q_norm)
                    else:
                        q_parts = query_value.split("-")
                        if len(q_parts) != 2:
                            return False
                        q_first = q_parts[0].strip()
                        q_second = str(int(q_parts[1]))
                        m = re.fullmatch(r'\s*(\d+)\s*-\s*(\d+)\s*', candidate_value)
                        if not m:
                            return False
                        cand_first, cand_second = m.groups()
                        return cand_first.startswith(q_first) and (cand_second == q_second)
                except Exception:
                    return False

            def coincide_caso(row):
                campo_origen = row.get("origen") or ""
                segmentos_origen = [seg.strip() for seg in campo_origen.split(",")]
                valid_origen = []
                for seg in segmentos_origen:
                    if "CASO" in seg.upper():
                        m = re.search(r'(?i)CASO\s*([\d-]+)', seg)
                        if m:
                            valid_origen.append(m.group(1).strip().upper())
                match_origen = any(
                    partial_numeric_match_normal(seg, query_sin_prefijo) or 
                    partial_numeric_match_inverted(seg, query_invertida)
                    for seg in valid_origen
                )
                campo_exp = row.get("nr de exp completo") or ""
                segmentos_exp = [seg.strip() for seg in campo_exp.split(",")]
                valid_exp = []
                for seg in segmentos_exp:
                    if re.fullmatch(patron_exp_flexible, seg):
                        partes_seg = [p.strip() for p in seg.split("-")]
                        if len(partes_seg) >= 4:
                            relevante = f"{partes_seg[-3]}-{partes_seg[-2]}".upper()
                        elif len(partes_seg) == 3:
                            relevante = f"{partes_seg[-2]}-{partes_seg[-1]}".upper()
                        else:
                            continue
                        valid_exp.append(relevante)
                match_exp = any(
                    partial_numeric_match_normal(seg, query_sin_prefijo) or 
                    partial_numeric_match_inverted(seg, query_invertida)
                    for seg in valid_exp
                )
                return match_origen or match_exp

            filtrados = [r for r in candidatos if coincide_caso(r)]
            resultados_unicos = []
            vistos = set()
            for r in filtrados:
                clave = r.get("id")
                if clave is None:
                    clave = str(r)
                if clave not in vistos:
                    vistos.add(clave)
                    resultados_unicos.append(r)
            combined_results = resultados_unicos[:10]
            return jsonify(combined_results), 200

        # Rama para el campo "casoJudicial"
        elif search_field == "casoJudicial":
            sql_datapenal = """
                SELECT
                    *,
                    LENGTH(origen) AS match_length,
                    'datapenal' AS source
                FROM datapenal
                WHERE origen REGEXP %s OR origen REGEXP %s
            """
            params_datapenal = (exp_pattern_1, exp_pattern_2)
            cursor.execute(sql_datapenal, params_datapenal)
            results_datapenal = cursor.fetchall()
            sql_consulta = """
                SELECT
                    *,
                    LENGTH(origen) AS match_length,
                    'consulta' AS source
                FROM consulta_ppupenal
                WHERE origen REGEXP %s OR origen REGEXP %s
            """
            params_consulta = (exp_pattern_1, exp_pattern_2)
            cursor.execute(sql_consulta, params_consulta)
            results_consulta = cursor.fetchall()
            candidatos = results_datapenal + results_consulta

            def coincide_caso_judicial(row):
                exp_full = (row.get("origen") or "").replace("Exp.", "").replace("CASO", "").strip()
                exp_parts = exp_full.split("-")
                if not exp_parts:
                    return False
                query_clean = query.strip().upper()
                query_fragments = query_clean.split("-")
                if not query_fragments:
                    return False
                if query_clean.endswith("-"):
                    if query_fragments[0].isdigit():
                        query_first = query_fragments[0].zfill(5)
                    else:
                        return False
                    return exp_parts[0] == query_first
                exp_first = exp_parts[0]
                query_first = query_fragments[0]
                if query_first.startswith("0"):
                    if not exp_first.upper().startswith(query_first):
                        return False
                else:
                    if not exp_first.lstrip("0").startswith(query_first):
                        return False
                for frag in query_fragments[1:]:
                    if not any(frag in part for part in exp_parts[1:]):
                        return False
                return True

            filtrados = [r for r in candidatos if coincide_caso_judicial(r)]
            combined_results = filtrados[:10]

        # Rama para el campo "denunciado"
        elif search_field == "denunciado":
            # Filtro preliminar: limitar la cantidad de registros solicitados a la base de datos.
            preliminary_query = f"%{query.upper()}%"
            sql_datapenal = """
                SELECT
                    *,
                    'datapenal' AS source
                FROM datapenal
                WHERE denunciado IS NOT NULL AND UPPER(denunciado) LIKE %s
                LIMIT 100
            """
            cursor.execute(sql_datapenal, (preliminary_query,))
            results_datapenal = cursor.fetchall()

            sql_consulta = """
                SELECT
                    *,
                    'consulta' AS source
                FROM consulta_ppupenal
                WHERE denunciado IS NOT NULL AND UPPER(denunciado) LIKE %s
                LIMIT 100
            """
            cursor.execute(sql_consulta, (preliminary_query,))
            results_consulta = cursor.fetchall()

            candidates = results_datapenal + results_consulta

            # Se intenta importar una versión optimizada de Levenshtein (rapidfuzz).
            try:
                from rapidfuzz.distance.Levenshtein import distance as levenshtein_distance
            except ImportError:
                # Si no está disponible, se utiliza la implementación propia en Python.
                def levenshtein_distance(s, t):
                    if s == t:
                        return 0
                    if len(s) == 0:
                        return len(t)
                    if len(t) == 0:
                        return len(s)
                    v0 = list(range(len(t) + 1))
                    for i in range(1, len(s) + 1):
                        v1 = [i] + [0] * len(t)
                        for j in range(1, len(t) + 1):
                            cost = 0 if s[i - 1] == t[j - 1] else 1
                            v1[j] = min(v1[j - 1] + 1, v0[j] + 1, v0[j - 1] + cost)
                        v0 = v1
                    return v0[len(t)]

            def normalize_text(text):
                text = text.upper().strip()
                replacements = {
                    "V": "B",
                    "Z": "S",
                    "Y": "I"
                }
                for old, new in replacements.items():
                    text = text.replace(old, new)
                return text

            normalized_query = normalize_text(query)

            filtered = []
            for row in candidates:
                name = row.get("denunciado")
                if name:
                    # Se separa el campo en tokens, usando la coma como delimitador.
                    tokens = [token.strip() for token in name.split(",")]
                    min_distance = float('inf')
                    for token in tokens:
                        norm_token = normalize_text(token)
                        # Se verifica si el término de búsqueda ya está contenido (coincidencia parcial).
                        if normalized_query in norm_token:
                            dist = 0
                        else:
                            dist = levenshtein_distance(normalized_query, norm_token)
                        if dist < min_distance:
                            min_distance = dist
                        if min_distance == 0:
                            break  # No es necesario seguir evaluando si se encontró una coincidencia exacta parcial.
                    threshold = 2
                    if min_distance <= threshold:
                        row["levenshtein"] = min_distance
                        filtered.append(row)

            filtered.sort(key=lambda x: x["levenshtein"])
            combined_results = filtered[:10]



        # Rama genérica para otros campos
        else:
            partial_pattern = f"%{query}%"
            where_clauses_datapenal = [f"{field_map_datapenal[search_field]} LIKE %s"]
            params_datapenal_list = [partial_pattern]
            if re.search(exp_pattern_1, query) or re.search(exp_pattern_2, query):
                normalized = normalizar_expediente(query)
                where_clauses_datapenal.append(f"{field_map_datapenal[search_field]} LIKE %s")
                params_datapenal_list.append(f"%{normalized}%")
            sql_datapenal = f"""
                SELECT
                    *,
                    LENGTH({field_map_datapenal[search_field]}) AS match_length,
                    'datapenal' AS source
                FROM datapenal
                WHERE {' OR '.join(where_clauses_datapenal)}
                ORDER BY match_length ASC
                LIMIT 10
            """
            params_datapenal = tuple(params_datapenal_list)
            cursor.execute(sql_datapenal, params_datapenal)
            results_datapenal = cursor.fetchall()
            where_clauses_consulta = [f"{field_map_consulta[search_field]} LIKE %s"]
            params_consulta_list = [partial_pattern]
            if re.search(exp_pattern_1, query) or re.search(exp_pattern_2, query):
                normalized = normalizar_expediente(query)
                where_clauses_consulta.append(f"{field_map_consulta[search_field]} LIKE %s")
                params_consulta_list.append(f"%{normalized}%")
            sql_consulta = f"""
                SELECT
                    *,
                    LENGTH({field_map_consulta[search_field]}) AS match_length,
                    'consulta' AS source
                FROM consulta_ppupenal
                WHERE {' OR '.join(where_clauses_consulta)}
                ORDER BY match_length ASC
                LIMIT 10
            """
            params_consulta = tuple(params_consulta_list)
            cursor.execute(sql_consulta, params_consulta)
            results_consulta = cursor.fetchall()
            combined_results = results_datapenal + results_consulta

        return jsonify(combined_results), 200

    except Exception as e:
        logger.error(f"Error en new_search: {e}", exc_info=True)
        return jsonify({"error": "Error interno del servidor"}), 500

    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()












################################################################################################################################################
### MODO BUSQUEDA TAB 0 ###########################################################################################################################
################################################################################################################################################

#####################MODO IMPULSO######################################################################################################################
from PyPDF2 import PdfMerger  # Se requiere instalar PyPDF2
from flask import request, jsonify
from werkzeug.utils import secure_filename

# ---------------- IMPULSO UPLOAD (fragmento completo) ---------------- #
@app.route('/api/impulso/upload', methods=['POST'])
@login_required
def impulso_upload():
    """
    Sube 1‑4 PDF, fusiona si es necesario y registra la acción de IMPULSO
    en datapenal_versioning y datapenal_plazos.
    El campo e_situacional se guarda como:
        IMPULSO (<fecha de hoy>): <accion>
    """
    # ---------- Validaciones ----------
    files = request.files.getlist('pdfs')
    if not files:
        return jsonify({"error": "No se encontraron archivos PDF"}), 400
    if len(files) > 4:
        return jsonify({"error": "Se permite un máximo de 4 archivos PDF"}), 400

    accion_valor = request.form.get("accion", "").strip()
    registro_ppu = request.form.get("registro_ppu", "").strip()
    origen       = request.form.get("origen", "").strip() or "SIN_ORIGEN"
    if not accion_valor or not registro_ppu:
        return jsonify({"error": "Los campos 'accion' y 'registro_ppu' son requeridos"}), 400

    # ---------- Campos opcionales ----------
    abogado            = request.form.get("abogado", "").strip()
    denunciado         = request.form.get("denunciado", "").strip()
    juzgado            = request.form.get("juzgado", "").strip()
    fiscalia           = request.form.get("fiscalia", "").strip()
    departamento       = request.form.get("departamento", "").strip()
    nr_de_exp_completo = request.form.get("nr_de_exp_completo", "").strip()
    delito             = request.form.get("delito", "").strip()
    informe_juridico   = request.form.get("informe_juridico", "").strip()
    item               = request.form.get("item", "").strip()

    # ---------- Fecha y carpeta destino ----------
    from datetime import datetime
    today        = datetime.now()
    fecha_hoy    = today.strftime("%d-%m-%Y")                 # <- para e_situacional
    months_map   = {1:"ENERO",2:"FEBRERO",3:"MARZO",4:"ABRIL",5:"MAYO",6:"JUNIO",
                    7:"JULIO",8:"AGOSTO",9:"SETIEMBRE",10:"OCTUBRE",11:"NOVIEMBRE",12:"DICIEMBRE"}
    dest_folder  = os.path.join(
        DESTINATION_BASE_PATH,
        str(today.year),
        months_map[today.month],
        fecha_hoy
    )
    os.makedirs(dest_folder, exist_ok=True)

    final_filename = f"IMPULSO {abogado} {registro_ppu} {origen}.pdf".strip()
    final_path     = os.path.join(dest_folder, final_filename)

    # ---------- Guardar / fusionar PDF ----------
    if len(files) > 1:
        merger = PdfMerger()
        for f in files:
            if not allowed_file(f.filename):
                return jsonify({"error": f"Archivo {f.filename} no permitido"}), 400
            f.stream.seek(0)
            merger.append(f.stream)
        merger.write(final_path)
        merger.close()
    else:
        f = files[0]
        if not allowed_file(f.filename):
            return jsonify({"error": f"Archivo {f.filename} no permitido"}), 400
        f.save(final_path)

    # ---------- Hash ----------
    import hashlib
    sha256 = hashlib.sha256()
    with open(final_path, "rb") as pdf:
        for chunk in iter(lambda: pdf.read(4096), b""):
            sha256.update(chunk)
    file_hash = sha256.hexdigest()

    # ---------- e_situacional automático ----------
    e_situacional_impulso = f"IMPULSO ({fecha_hoy}): {accion_valor}"

    # ---------- Inserciones en BD ----------
    conn = get_db_connection()
    if not conn:
        return jsonify({"error": "Error al conectar con la base de datos"}), 500
    try:
        cur = conn.cursor()

        # ---- datapenal_versioning ----
        sql_versioning = """
            INSERT INTO datapenal_versioning
              (abogado, registro_ppu, denunciado, origen, `nr de exp completo`, delito,
               informe_juridico, item, juzgado, fiscalia, departamento, e_situacional,
               fecha_version, usuario_modificacion, fecha_e_situacional,
               ruta, hash_sha, clasificacion, consulta_ppu)
            VALUES
              (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
               NOW(),%s,CURDATE(),%s,%s,'','')
        """
        cur.execute(
            sql_versioning,
            (abogado, registro_ppu, denunciado, origen, nr_de_exp_completo, delito,
             informe_juridico, item, juzgado, fiscalia, departamento, e_situacional_impulso,
             session.get("username",""), final_path, file_hash)
        )

        # ---- datapenal_plazos ----
        accion_impulso = f"IMPULSO - {accion_valor}"
        sql_plazos = """
            INSERT INTO datapenal_plazos
              (abogado, registro_ppu, denunciado, origen, `nr de exp completo`, delito,
               informe_juridico, item, juzgado, fiscalia, departamento, e_situacional,
               fecha_ingreso,      last_modified, etiqueta, accion,
               plazo_atencion,     fecha_atencion, seguimiento,
               ruta, hash_sha, ruta_respuesta, hash_respuesta,
               consulta_ppu, tipoPlazo)
            VALUES
              (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
               NULL,              NOW(),        '',      %s,
               'NA',              NOW(),        'ATENDIDA',
               %s,%s,'','',        '',           'IMPULSO')
        """
        cur.execute(
            sql_plazos,
            (abogado, registro_ppu, denunciado, origen, nr_de_exp_completo, delito,
             informe_juridico, item, juzgado, fiscalia, departamento, e_situacional_impulso,
             accion_impulso, final_path, file_hash)
        )

        conn.commit()
        return jsonify({"message": "Operación completada"}), 200

    except Exception as e:
        conn.rollback()
        logger.error("Error en impulso upload: %s", e, exc_info=True)
        return jsonify({"error": f"Error en la operación: {e}"}), 500
    finally:
        cur.close()
        conn.close()
# -------------------------------------------------------------------- #

# -------------------------------------------------------------------- #
# ---------------------------------OBSERVACION ----------------------- #
# -------------------------------------------------------------------- #


@app.route('/api/guardar_observacion', methods=['POST'])
@login_required  # Opcional: si quieres exigir que el usuario esté logueado
def guardar_observacion():
    """
    Este endpoint recibe un JSON con:
      - id:        el id de la fila en datapenal_plazos (entero o string convertible a entero)
      - observacion: texto ingresado por el abogado en el modal

    Se encarga de:
      1. Validar que llegue el campo 'id'.
      2. Si 'observacion' queda vacío o sólo espacios:
           • poner observacion_abogado = ''  
           • si seguimiento estaba en 'OBSERVADO', dejarlo en ''
      3. Si 'observacion' contiene texto válido:
           • poner observacion_abogado = <valor>
           • seguimiento = 'OBSERVADO'
      4. Actualizar last_modified = NOW()
      5. Devolver un JSON con el resultado exitoso o mensaje de error.
    """
    app.logger.info("Entrando a endpoint /api/guardar_observacion")
    payload = request.json or {}
    app.logger.debug(f"Payload recibido: {payload}")

    fila_id = payload.get('id')
    observacion = payload.get('observacion', '')
    observacion_stripped = observacion.strip()
    app.logger.info(f"ID recibido: {fila_id}, Observación (bruta): '{observacion}'")

    # 1) Validación de 'id'
    if not fila_id:
        app.logger.warning("No se proporcionó 'id' en el payload.")
        return jsonify({"error": "Se requiere el parámetro 'id'."}), 400

    try:
        # 2) Conexión a la base de datos
        app.logger.info("Intentando conectar a la base de datos...")
        conn = get_db_connection()
        if conn is None:
            app.logger.error("get_db_connection() devolvió None.")
            return jsonify({"error": "Error al conectar con la base de datos."}), 500

        cursor = conn.cursor()
        app.logger.info("Conexión a la base de datos exitosa. Obteniendo cursor.")

        # 3) Decidir comportamiento según si quedó vacío o no
        if observacion_stripped == '':
            app.logger.info("Observación vacía o sólo espacios. Preparando SQL para borrar observación.")
            sql_update = """
                UPDATE datapenal_plazos
                SET
                    observacion_abogado = '',
                    seguimiento         = '',
                    last_modified       = NOW()
                WHERE id = %s
            """
            params = (fila_id,)
            app.logger.debug(f"SQL a ejecutar (vaciar campos): {sql_update.strip()}, params={params}")
        else:
            app.logger.info("Observación válida proporcionada. Preparando SQL para guardar observación.")
            sql_update = """
                UPDATE datapenal_plazos
                SET 
                    observacion_abogado = %s,
                    seguimiento         = %s,
                    last_modified       = NOW()
                WHERE id = %s
            """
            params = (observacion_stripped, 'OBSERVADO', fila_id)
            app.logger.debug(f"SQL a ejecutar (guardar campos): {sql_update.strip()}, params={params}")

        cursor.execute(sql_update, params)
        conn.commit()
        app.logger.info(f"Commit ejecutado. Filas afectadas: {cursor.rowcount}")

        # 4) Verificar si realmente se modificó alguna fila
        if cursor.rowcount == 0:
            app.logger.warning(f"No se encontró ningún registro con id = {fila_id}.")
            return jsonify({"error": f"No se encontró ningún registro con id = {fila_id}."}), 404

        # 5) Armar la respuesta
        if observacion_stripped == '':
            app.logger.info("Observación eliminada correctamente.")
            return jsonify({
                "message": "Observación borrada y seguimiento limpiado.",
                "id_actualizado": fila_id,
                "seguimiento": ""
            }), 200
        else:
            app.logger.info("Observación guardada correctamente y seguimiento marcado como 'OBSERVADO'.")
            return jsonify({
                "message": "Observación guardada correctamente.",
                "id_actualizado": fila_id,
                "seguimiento": "OBSERVADO"
            }), 200

    except mysql.connector.Error as db_err:
        # 6) Si ocurre algún error en MySQL
        app.logger.error(f"Error al actualizar observación en MySQL: {db_err}", exc_info=True)
        return jsonify({"error": "Error en la base de datos al guardar la observación."}), 500

    except Exception as exc:
        # Capturar cualquier otro error inesperado
        app.logger.error(f"Error inesperado en guardar_observacion: {exc}", exc_info=True)
        return jsonify({"error": "Ha ocurrido un error inesperado."}), 500

    finally:
        # 7) Cierre de cursor y conexión
        try:
            cursor.close()
            app.logger.debug("Cursor cerrado correctamente.")
        except Exception:
            app.logger.warning("No se pudo cerrar el cursor. Probablemente nunca se abrió.")
        try:
            conn.close()
            app.logger.debug("Conexión cerrada correctamente.")
        except Exception:
            app.logger.warning("No se pudo cerrar la conexión. Probablemente nunca se abrió.")




# -------------------------------------------------------------------- #
# ---------------------------------OBSERVACION ----------------------- #
# -------------------------------------------------------------------- #


# -------------------------------------------------------------------- #
# ---------------------------------MODO BUSQUEDA ----------------------- #
# -------------------------------------------------------------------- #

@app.route('/api/busqueda_rapida', methods=['GET'])
@login_required
def busqueda_rapida():
    q       = request.args.get('q', '').strip()
    origen  = request.args.get('origen', '').strip()
    depto   = request.args.get('departamento', '').strip()

    if not q:
        return jsonify([])

    pattern = f"%{q}%"
    conn    = get_db_connection()
    if conn is None:
        return jsonify([]), 500

    try:
        cursor = conn.cursor(dictionary=True)
        sql = """
         SELECT
            abogado                                   AS abogado,
            registro_ppu                              AS `registro_ppu`,
            registro_ppu                              AS registroPpu,
            denunciado                                AS denunciado,
            origen                                    AS origen,
            `nr de exp completo`                      AS nr_de_exp_completo,
            fiscalia                                  AS fiscaliaOrigen,
            departamento                              AS departamento,
            juzgado                                   AS juzgado,
            delito                                    AS delito,
            e_situacional                             AS e_situacional,
            informe_juridico                          AS informeJuridico,
            item                                      AS item,
            fecha_ingreso                             AS fechaIngreso,
            etiqueta                                  AS etiqueta,
            fecha_de_archivo                          AS fechaDeArchivo,
            razon_archivo                             AS razonArchivo
         FROM datapenal
         WHERE
            (registro_ppu         LIKE %s OR
             abogado              LIKE %s OR
             denunciado           LIKE %s OR
             origen               LIKE %s OR
             `nr de exp completo` LIKE %s OR
             fiscalia             LIKE %s OR
             departamento         LIKE %s OR
             juzgado              LIKE %s OR
             delito               LIKE %s OR
             e_situacional        LIKE %s OR
             informe_juridico     LIKE %s)
        """
        params = [pattern] * 11

        if origen:
            sql += " AND origen LIKE %s"
            params.append(f"%{origen}%")
        if depto:
            sql += " AND departamento LIKE %s"
            params.append(f"%{depto}%")

        sql += " ORDER BY registro_ppu LIMIT 50"

        cursor.execute(sql, params)
        rows = cursor.fetchall()

        # Normaliza fechas a ISO para el frontend (keys camelCase que la grilla espera)
        from datetime import date, datetime
        for r in rows:
            if r.get('fechaIngreso') and isinstance(r['fechaIngreso'], (date, datetime)):
                r['fechaIngreso'] = r['fechaIngreso'].strftime('%Y-%m-%d')
            if r.get('fechaDeArchivo') and isinstance(r['fechaDeArchivo'], (date, datetime)):
                r['fechaDeArchivo'] = r['fechaDeArchivo'].strftime('%Y-%m-%d')

        return jsonify(rows)
    except Exception:
        return jsonify([]), 500
    finally:
        cursor.close()
        conn.close()



@app.route('/api/juzgado_incompleto', methods=['GET'])
@login_required
def juzgado_incompleto():
    pattern = request.args.get('pattern', '').strip()
    if not pattern:
        return jsonify([]), 400

    # Conecta a la BD monitoreo_descargas_sinoe
    conn = get_db_connection(database='monitoreo_descargas_sinoe')
    if conn is None:
        return jsonify([]), 500

    try:
        cursor = conn.cursor(dictionary=True)
        sql = """
          SELECT juzgado_incompleto
          FROM conteo_exp
          WHERE nombre_original REGEXP %s
          LIMIT 1
        """
        cursor.execute(sql, (pattern,))
        row = cursor.fetchone()
        return jsonify([row] if row else [])
    except Exception as e:
        app.logger.error("juzgado_incompleto error:", exc_info=True)
        return jsonify([]), 500
    finally:
        cursor.close()
        conn.close()

@app.route('/api/fiscalia_incompleto', methods=['GET'])
@login_required
def fiscalia_incompleto():
    pattern = request.args.get('pattern', '').strip()
    if not pattern:
        return jsonify([]), 400

    # Conecta a la BD que contiene la tabla dependencias_fiscales_mpfn
    conn = get_db_connection(database='datappupenal')
    if conn is None:
        return jsonify([]), 500

    try:
        cursor = conn.cursor(dictionary=True)
        # Buscamos por las primeras 10 cifras exactas
        regexp = f'^{pattern}'
        sql = """
          SELECT nr_de_exp_completo, fiscalia, departamento
          FROM dependencias_fiscales_mpfn
          WHERE nr_de_exp_completo REGEXP %s
          LIMIT 1
        """
        cursor.execute(sql, (regexp,))
        row = cursor.fetchone()
        return jsonify([row] if row else [])
    except Exception:
        app.logger.error("fiscalia_incompleto error:", exc_info=True)
        return jsonify([]), 500
    finally:
        cursor.close()
        conn.close()


# Campos que puede enviar el frontend (claves del payload)
ALL_FRONT_FIELDS = [
    # claves usadas por el front / existentes en datapenal
    'registroPpu',
    'abogado', 'denunciado', 'origen',
    'nrDeExpCompleto',      # ← alias camel de `nr de exp completo`
    'fiscaliaOrigen',       # ← alias del campo fiscalia
    'departamento', 'juzgado', 'delito',
    'informeJuridico', 'item',
    'eSituacional',         # ← alias camel de e_situacional
    'fechaIngreso',
    'etiqueta',
    'fechaDeArchivo',       # ← nuevo: existe en la tabla
    'razonArchivo',  
]


def get_allowed_fields_for_user(username: str, users: dict) -> set:
    """
    Política de edición por usuario:
      - No admins: nada (solo ver).
      - Manuel (admin): todo.
      - agarcia (admin): nada (bloqueado).
      - jgranda (admin): todo MENOS eSituacional.
      - Resto de admins: todo.
    """
    info = users.get(username, {})
    role = info.get("role", "user")

    if role != "admin":
        return set()  # no admins no pueden editar nada

    if username == "Manuel":
        return set(ALL_FRONT_FIELDS)

    if username == "agarcia":
        return set()  # admin sin permisos de edición

    if username == "jgranda":
        return set([f for f in ALL_FRONT_FIELDS if f != 'eSituacional'])

    # cualquier otro admin
    return set(ALL_FRONT_FIELDS)

@app.route('/api/me', methods=['GET'])
@login_required
def api_me():
    """
    Devuelve identidad, permisos y si el usuario puede editar algo.
    """
    username = session.get('username') or ''
    role = session.get('role') or 'user'

    allowed_fields = get_allowed_fields_for_user(username, users)
    allowed_list = sorted(allowed_fields)

    return jsonify({
        "username": username,
        "role": role,
        "allowedFields": allowed_list,  # Lista de campos que puede editar
        "canEdit": bool(allowed_list)   # True si hay al menos un campo editable
    }), 200




@app.route('/api/busqueda_rapida_sync', methods=['POST'])
@login_required
def busqueda_rapida_sync():
    username = session.get("username")
    if not username:
        return jsonify(error="No autenticado"), 401

    allowed_fields = get_allowed_fields_for_user(username, users)
    if not allowed_fields:
        app.logger.warning(f"busqueda_rapida_sync: Acceso denegado para {username}")
        return jsonify(error="Acceso denegado"), 403

    rows = request.get_json() or []
    if not isinstance(rows, list):
        app.logger.warning('busqueda_rapida_sync: payload inválido, se esperaba lista de registros')
        return jsonify(updated=[]), 400

    conn = get_db_connection()
    if conn is None:
        app.logger.error('busqueda_rapida_sync: error al conectar a la base de datos')
        return jsonify(updated=[]), 500

    cursor = conn.cursor(dictionary=True)

    # Alinea el NOW() del trigger con hora de Lima (UTC-5)
    try:
        cursor.execute("SET time_zone = '-05:00'")
    except Exception:
        app.logger.warning("No se pudo fijar time_zone de sesión; se continúa con la del servidor.")

    updated_ppus = []

    # Normaliza cualquier tipo a string “segura” para comparar (no para escribir fechas)
    def _s(v):
        if v is None:
            return ''
        try:
            return str(v).strip()
        except Exception:
            return ''

    # Normaliza fechas a formato aceptado por MySQL o None (NULL)
    # Acepta: '', None, 'YYYY-MM-DD', 'YYYY-MM-DD HH:MM:SS', 'DD/MM/YYYY', 'DD-MM-YYYY'
    from datetime import datetime, date
    def _date_norm(v):
        if v is None:
            return None
        if isinstance(v, (datetime,)):
            return v.strftime('%Y-%m-%d')
        if isinstance(v, (date,)):
            return v.strftime('%Y-%m-%d')
        sv = str(v).strip()
        if not sv:
            return None
        for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
            try:
                dt = datetime.strptime(sv, fmt)
                return dt.strftime('%Y-%m-%d')
            except ValueError:
                continue
        # Si no parsea, mejor enviar NULL que cadena vacía
        return None

    try:
        for idx, new_row in enumerate(rows, start=1):
            ppu = new_row.get('registroPpu')
            if not ppu:
                app.logger.warning(f"[Fila {idx}] Sin 'registroPpu', se omite")
                continue

            # 🔒 Defensa: ignorar cualquier intento de setear fechaDeArchivo desde el front
            if 'fechaDeArchivo' in new_row:
                new_row.pop('fechaDeArchivo', None)

            app.logger.info(f"[Fila {idx}] ({username}) Datos recibidos (PPU: {ppu}): {new_row}")

            # 1) Traer valores actuales desde la base (todas las columnas que se comparan)
            cursor.execute("""
                SELECT
                    abogado,
                    denunciado,
                    origen,
                    `nr de exp completo`  AS nr_de_exp_completo,
                    fiscalia              AS fiscalia_origen,
                    departamento,
                    juzgado,
                    delito,
                    e_situacional,
                    informe_juridico,
                    item,
                    fecha_ingreso,
                    fecha_e_situacional,
                    etiqueta,
                    last_modified,
                    fecha_de_archivo,
                    razon_archivo AS razonArchivo
                FROM datapenal
                WHERE registro_ppu = %s
            """, (ppu,))
            db_row = cursor.fetchone()
            if not db_row:
                app.logger.warning(f"[Fila {idx}] No se encontró registro en BD para PPU: {ppu}")
                continue

            # 2) Comparar campo a campo (solo los permitidos)
            diffs = {}

            # ⚠️ Quitamos 'fechaDeArchivo' del comparador. La maneja el TRIGGER en BD.
            comparadores = [
                ('abogado',            'abogado'),
                ('denunciado',         'denunciado'),
                ('origen',             'origen'),
                ('nrDeExpCompleto',    'nr_de_exp_completo'),
                ('fiscaliaOrigen',     'fiscalia_origen'),
                ('departamento',       'departamento'),
                ('juzgado',            'juzgado'),
                ('delito',             'delito'),
                ('informeJuridico',    'informe_juridico'),
                ('item',               'item'),
                ('eSituacional',       'e_situacional'),
                ('fechaIngreso',       'fecha_ingreso'),
                ('fechaESituacional',  'fecha_e_situacional'),
                ('etiqueta',           'etiqueta'),
                ('razonArchivo',       'razon_archivo')
            ]

            # Solo fechas que el front puede modificar (NO incluir fecha_de_archivo ni last_modified)
            date_fields = {'fecha_ingreso', 'fecha_e_situacional'}

            for front_key, db_key in comparadores:
                if front_key not in allowed_fields:
                    continue

                if db_key in date_fields:
                    new_val_norm = _date_norm(new_row.get(front_key))
                    old_raw = db_row.get(db_key)

                    if isinstance(old_raw, (datetime, date)):
                        old_val_norm = old_raw.strftime('%Y-%m-%d')
                    else:
                        old_val_norm = _date_norm(old_raw)

                    if new_val_norm != old_val_norm:
                        if db_key == 'nr_de_exp_completo':
                            diffs['`nr de exp completo`'] = new_val_norm
                        elif db_key == 'fiscalia_origen':
                            diffs['fiscalia'] = new_val_norm
                        else:
                            diffs[db_key] = new_val_norm  # None → NULL
                else:
                    new_val = _s(new_row.get(front_key))
                    old_val = _s(db_row.get(db_key))
                    if new_val != old_val:
                        if db_key == 'nr_de_exp_completo':
                            diffs['`nr de exp completo`'] = new_val
                        elif db_key == 'fiscalia_origen':
                            diffs['fiscalia'] = new_val
                        else:
                            diffs[db_key] = new_val

            # 3) Si hay diferencias permitidas, actualizar
            if diffs:
                # 🔒 Defensa extra: por si apareció por error
                diffs.pop('fecha_de_archivo', None)

                set_parts = [f"{col} = %s" for col in diffs.keys()]
                params = list(diffs.values()) + [ppu]
                sql = f"UPDATE datapenal SET {', '.join(set_parts)} WHERE registro_ppu = %s"
                cursor.execute(sql, params)

                updated_ppus.append(ppu)
                app.logger.info(f"[Fila {idx}] ({username}) Cambios aplicados PPU {ppu}: {list(diffs.keys())}")
            else:
                app.logger.info(f"[Fila {idx}] ({username}) Sin cambios (o sin permisos) para PPU: {ppu}")

        conn.commit()
        app.logger.info(f"busqueda_rapida_sync: total registros actualizados → {len(updated_ppus)} por {username}")
        return jsonify(updated=updated_ppus), 200

    except Exception:
        conn.rollback()
        app.logger.exception('busqueda_rapida_sync error', exc_info=True)
        return jsonify(updated=[]), 500
    finally:
        try:
            cursor.close()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass


# -------------------------HISTORIAL BOTON--------------- #



# columnas que el grid envía (snakeCase)  ⇢  columna real en BD
# columnas que existen en AMBAS tablas (datapenal y datapenal_versioning)
COLUMNS_MAP = {
    'registro_ppu':           '`registro_ppu`',
    'abogado':                '`abogado`',
    'denunciado':             '`denunciado`',
    'origen':                 '`origen`',
    'nr_de_exp_completo':     '`nr de exp completo`',
    'fiscaliaOrigen':         '`fiscalia`',
    'fiscalia':               '`fiscalia`',
    'departamento':           '`departamento`',
    'juzgado':                '`juzgado`',
    'delito':                 '`delito`',
    'e_situacional':          '`e_situacional`',
    'informe_juridico':       '`informe_juridico`',
    'item':                   '`item`',
    'fecha_ingreso':          '`fecha_ingreso`',
    'etiqueta':               '`etiqueta`',
    'fecha_de_archivo':       '`fecha_de_archivo`',
    'razon_archivo':          '`razon_archivo`',
}

# columnas que existen SOLO en datapenal_versioning
VERSION_ONLY_COLUMNS = {
    'ruta': '`ruta`',
}



def _norm(v):
    """igual que en laboral: quita mayúsc/minúsc-tildes-espacios para comparar."""
    import unidecode, re
    if v is None:
        return ''
    v = str(v).strip().lower()
    v = unidecode.unidecode(v)
    v = re.sub(r'\s+', ' ', v)
    return v

# 1) ¿QUÉ CAMPOS CAMBIARON? ──────────────────────────────────
@app.route('/api/busqueda_rapida_history_available', methods=['GET'])
@login_required
def history_available_penal():
    app.logger.debug(">> /history_available – inicio")
    ppu = request.args.get('ppu', '').strip()
    app.logger.debug("   parámetros → ppu='%s'", ppu)

    if not ppu:
        app.logger.warning("   ppu vacío; 400")
        return jsonify(success=False, fields=[]), 400

    conn = get_db_connection()          # BD «datappupenal»
    if conn is None:
        app.logger.error("   conexión BD fallida; 500")
        return jsonify(success=False, fields=[]), 500
    cur = conn.cursor(dictionary=True)

    try:
        # 1️⃣ versiones guardadas
        cols_versions = ",\n       ".join(
            [f"{col} AS {snake}" for snake, col in COLUMNS_MAP.items()] +
            [f"{col} AS {snake}" for snake, col in VERSION_ONLY_COLUMNS.items()]
        )

        sql_versions = f"""
            SELECT version_id,
                   {cols_versions}
              FROM datapenal_versioning
             WHERE registro_ppu = %s
             ORDER BY version_id
        """

        cols_current = ", ".join(f"{col} AS {snake}" for snake, col in COLUMNS_MAP.items())

        sql_current = f"""
            SELECT {cols_current}
              FROM datapenal
             WHERE registro_ppu = %s
        """

        app.logger.debug("   SQL versiones:\n%s", sql_versions)
        cur.execute(sql_versions, (ppu,))
        versions = cur.fetchall()
        app.logger.debug("   versiones encontradas: %d", len(versions))

        if not versions:
            return jsonify(success=True, fields=[]), 200

        # 2️⃣ fila actual
        sql_current = f"""
            SELECT {', '.join(f"{col} AS {snake}"
                              for snake, col in COLUMNS_MAP.items())}
              FROM datapenal
             WHERE registro_ppu = %s
        """
        cur.execute(sql_current, (ppu,))
        current = cur.fetchone() or {}
        app.logger.debug("   fila actual obtenida")

        # 3️⃣ comparar
        changed = []
        for snake in COLUMNS_MAP:
            cur_norm   = _norm(current.get(snake))
            hist_norms = [_norm(v[snake]) for v in versions
                          if v[snake] not in (None, '')]
            if any(h != cur_norm for h in hist_norms):
                changed.append(snake)

        app.logger.debug("   campos cambiados: %s", changed)
        return jsonify(success=True, fields=changed), 200

    except Exception:
        import traceback; traceback.print_exc()
        app.logger.exception("   ERROR en /history_available")
        return jsonify(success=False, fields=[]), 500

    finally:
        cur.close(); conn.close()
        app.logger.debug("<< /history_available – fin")

# ────────────────────────────────────────────────────────────
#  1-bis)  ¿QUÉ CAMPOS CAMBIARON?  (varios PPU a la vez)
#          POST  /api/busqueda_rapida_history_available_bulk
#          Body →  ["D-123-2024", "LEG-99-2022", …]
#          Res  →  {
#                     "D-123-2024": ["abogado","nr_de_exp_completo"],
#                     "LEG-99-2022": [],
#                     …
#                   }
# ────────────────────────────────────────────────────────────
@app.route('/api/busqueda_rapida_history_available_bulk', methods=['POST'])
@login_required
def history_available_bulk_penal():
    app.logger.debug(">> /history_available_bulk – inicio")

    ppus = request.get_json(silent=True) or []
    if not isinstance(ppus, list) or not ppus:
        app.logger.warning("   payload vacío o no-lista; 400")
        return jsonify({}), 400

    # ­­­— normaliza y elimina duplicados —­­­
    ppus = sorted({str(p).strip() for p in ppus if str(p).strip()})
    placeholders = ",".join(["%s"] * len(ppus))
    app.logger.debug("   ppus recibidos: %s", ppus)

    conn = get_db_connection()
    if conn is None:
        app.logger.error("   conexión BD fallida; 500")
        return jsonify({}), 500
    cur = conn.cursor(dictionary=True)

    try:
        # 1️⃣  Traer todas las versiones de TODOS los ppu solicitados
        cols_versions = ",\n       ".join(
            [f"{col} AS {snake}" for snake, col in COLUMNS_MAP.items()] +
            [f"{col} AS {snake}" for snake, col in VERSION_ONLY_COLUMNS.items()]
        )

        sql_versions = f"""
            SELECT registro_ppu,
                   version_id,
                   {cols_versions}
            FROM datapenal_versioning
            WHERE registro_ppu IN ({placeholders})
            ORDER BY registro_ppu, version_id
        """

        cols_current = ", ".join(f"{col} AS {snake}" for snake, col in COLUMNS_MAP.items())

        sql_current = f"""
            SELECT registro_ppu,
                   {cols_current}
            FROM datapenal
            WHERE registro_ppu IN ({placeholders})
        """

        app.logger.debug("   SQL versiones bulk:\n%s", sql_versions)
        cur.execute(sql_versions, ppus)
        all_versions = cur.fetchall()           # ← lista de dicts

        # 2️⃣  Fila actual para cada PPU
        sql_current = f"""
            SELECT registro_ppu,
                   {', '.join(f"{col} AS {snake}"
                              for snake, col in COLUMNS_MAP.items())}
            FROM datapenal
            WHERE registro_ppu IN ({placeholders})
        """
        app.logger.debug("   SQL current bulk:\n%s", sql_current)
        cur.execute(sql_current, ppus)
        current_rows = cur.fetchall()

        # ­­­—  dicts rápidos:  ppu -> filaActual  /  ppu -> [versions…] —­­­
        current_map  = {r['registro_ppu']: r for r in current_rows}
        versions_map = {}
        for v in all_versions:
            versions_map.setdefault(v['registro_ppu'], []).append(v)

        # 3️⃣  Comparar por cada ppu
        result = {}
        for ppu in ppus:
            cur_row   = current_map.get(ppu, {})
            versions  = versions_map.get(ppu, [])

            if not versions:
                result[ppu] = []
                continue

            changed = []
            for snake in COLUMNS_MAP:
                cur_norm   = _norm(cur_row.get(snake))
                hist_norms = [_norm(v[snake])
                              for v in versions
                              if v[snake] not in (None, '')]
                if any(h != cur_norm for h in hist_norms):
                    changed.append(snake)

            result[ppu] = changed

        app.logger.debug("   resultado bulk listo")
        return jsonify(result), 200

    except Exception:
        import traceback; traceback.print_exc()
        app.logger.exception("   ERROR en /history_available_bulk")
        return jsonify({}), 500

    finally:
        cur.close(); conn.close()
        app.logger.debug("<< /history_available_bulk – fin")

# 2) DETALLE DE UN CAMPO ─────────────────────────────────────
@app.route('/api/busqueda_rapida_history', methods=['GET'])
@login_required
def busqueda_rapida_history_penal():
    app.logger.debug(">> /history_detail – inicio")
    ppu   = request.args.get('ppu', '').strip()
    field = request.args.get('field', '').strip()
    app.logger.debug("   parámetros → ppu='%s', field='%s'", ppu, field)

    if not ppu or not field:
        app.logger.warning("   ppu o field faltantes; 400")
        return jsonify(success=False, message='ppu y field son obligatorios'), 400

    real_col = COLUMNS_MAP.get(field)
    if not real_col:
        app.logger.warning("   field desconocido: %s", field)
        return jsonify(success=False, message='Campo desconocido'), 400

    conn = get_db_connection()
    if conn is None:
        app.logger.error("   conexión BD fallida; 500")
        return jsonify(success=False, data=[]), 500
    cur = conn.cursor(dictionary=True)

    try:
        # 1) Valor actual (tabla principal)
        sql_current = f"""
            SELECT {real_col} AS cur_value
              FROM datapenal
             WHERE registro_ppu = %s
        """
        cur.execute(sql_current, (ppu,))
        cur_val = (cur.fetchone() or {}).get('cur_value')
        app.logger.debug("   valor actual = %s", cur_val)

        # 2) Versiones del campo (para e_situacional también traemos ruta)
        if field == 'e_situacional':
            sql_versions = f"""
                SELECT version_id,
                       {real_col}        AS old_value,
                       fecha_version,
                       usuario_modificacion,
                       `ruta`            AS ruta
                  FROM datapenal_versioning
                 WHERE registro_ppu = %s
                   AND {real_col} IS NOT NULL
                   AND CAST({real_col} AS CHAR) <> ''
                 ORDER BY version_id
            """
        else:
            sql_versions = f"""
                SELECT version_id,
                       {real_col}        AS old_value,
                       fecha_version,
                       usuario_modificacion
                  FROM datapenal_versioning
                 WHERE registro_ppu = %s
                   AND {real_col} IS NOT NULL
                   AND CAST({real_col} AS CHAR) <> ''
                 ORDER BY version_id
            """

        cur.execute(sql_versions, (ppu,))
        versions = cur.fetchall()
        app.logger.debug("   versiones encontradas: %d", len(versions))

        # 3) Armar lista según campo
        cur_norm = _norm(cur_val)

        if field == 'e_situacional':
            # Diferentes al actual
            different = [v for v in versions if _norm(v['old_value']) != cur_norm]

            # Entre las IGUALES al actual, priorizar la que tenga ruta; si no, la más reciente
            same_list = [dict(v) for v in reversed(versions) if _norm(v['old_value']) == cur_norm]  # más nueva → más vieja
            chosen_same = next((v for v in same_list if v.get('ruta')), None) or (same_list[0] if same_list else None)

            # “Actual” sintético desde la tabla principal (fallback si no hay chosen_same)
            current_row = {
                'version_id': 0,
                'old_value': cur_val,
                'fecha_version': None,
                'usuario_modificacion': '(actual)',
                # 'ruta': None  # no se incluye si no es de versioning; el front tolera ausencia
            }

            rows = []
            # Siempre mostramos una fila “Actual” en e_situacional (con ruta si hay, si no la sintética)
            if chosen_same:
                chosen_same['usuario_modificacion'] = '(actual)'
                rows.append(chosen_same)
            else:
                rows.append(current_row)

            # Si hay diferencias, también las agregamos
            rows.extend(different)

            filtered = rows

        else:
            # Para campos ≠ e_situacional:
            different = [v for v in versions if _norm(v['old_value']) != cur_norm]

            # Si NO hay diferencias, no hay historial
            if not different:
                return jsonify(success=True, data=[]), 200

            # Incluir el ACTUAL desde la tabla principal (no de versioning)
            current_row = {
                'version_id': 0,
                'old_value': cur_val,
                'fecha_version': None,
                'usuario_modificacion': '(actual)'
            }

            filtered = [current_row] + different

        # 4) Deduplicación por valor normalizado
        #    - e_situacional: preferir el que tenga ruta; si empatan, el más reciente.
        #    - otros campos: primera aparición (comportamiento clásico).
        from datetime import datetime

        def _parse_dt(s):
            try:
                if isinstance(s, datetime):
                    return s
                if not s:
                    return None
                try:
                    return datetime.strptime(str(s), "%Y-%m-%d %H:%M:%S")
                except ValueError:
                    try:
                        return datetime.strptime(str(s), "%d-%m-%Y %H:%M:%S")
                    except ValueError:
                        return None
            except Exception:
                return None

        if field == 'e_situacional':
            by_val = {}  # ov_norm -> item elegido
            for v in filtered:
                ov_norm = _norm(v['old_value'])
                prev = by_val.get(ov_norm)
                if prev is None:
                    by_val[ov_norm] = v
                    continue

                prev_has = bool(prev.get('ruta'))
                v_has = bool(v.get('ruta'))

                # 1) Preferir el que tenga ruta
                if v_has and not prev_has:
                    by_val[ov_norm] = v
                    continue
                if prev_has and not v_has:
                    continue

                # 2) Si ambos (o ninguno) tienen ruta, preferir MÁS RECIENTE
                prev_dt = _parse_dt(prev.get('fecha_version'))
                v_dt = _parse_dt(v.get('fecha_version'))
                if prev_dt and v_dt:
                    if v_dt > prev_dt:
                        by_val[ov_norm] = v
                else:
                    # Respaldo por version_id
                    try:
                        if int(v.get('version_id') or 0) > int(prev.get('version_id') or 0):
                            by_val[ov_norm] = v
                    except Exception:
                        pass

            unique = list(by_val.values())

        else:
            seen, unique = set(), []
            for v in filtered:
                ov_norm = _norm(v['old_value'])
                if ov_norm not in seen:
                    seen.add(ov_norm)
                    unique.append(v)

        result = unique[:1] if len(unique) == 1 else unique
        app.logger.debug("   valores devueltos: %d", len(result))

        # 5) Formateo final
        from datetime import datetime as _dt, date as _date
        def _fmt(x):
            if isinstance(x, (_dt, _date)):
                return x.strftime('%d-%m-%Y %H:%M:%S')
            return '' if x is None else str(x)

        history = []
        for r in result:
            item = {
                'version_id':           r.get('version_id'),
                'old_value':            _fmt(r.get('old_value')),
                'fecha_version':        _fmt(r.get('fecha_version')),
                'usuario_modificacion': _fmt(r.get('usuario_modificacion')),
            }
            # Incluir ruta SOLO para e_situacional (si la trae desde versioning)
            if field == 'e_situacional' and 'ruta' in r:
                item['ruta'] = r.get('ruta')
            history.append(item)

        return jsonify(success=True, data=history), 200

    except Exception as e:
        import traceback; traceback.print_exc()
        app.logger.exception("   ERROR en /history_detail")
        return jsonify(success=False, message=str(e)), 500

    finally:
        try:
            cur.close()
        finally:
            conn.close()
        app.logger.debug("<< /history_detail – fin")


from flask import send_file, abort
import os

# Debe coincidir con la base real donde viven los PDFs
ALLOWED_BASE = r'\\agarciaf\NOTIFICACIONES RENIEC\MESA DE PARTES\PENAL\NOTIFICACIONES'

def _is_safe_path(path, base):
    # Normaliza y comprueba que la ruta esté dentro de la base permitida
    p = os.path.normpath(path)
    b = os.path.normpath(base)
    try:
        return os.path.commonpath([p, b]) == b
    except Exception:
        return False

@app.route('/api/open_pdf_by_ruta', methods=['GET'])
@login_required
def open_pdf_by_ruta():
    app.logger.info("[open_pdf_by_ruta] Cookie=%r", request.headers.get('Cookie'))
    app.logger.info("[open_pdf_by_ruta] session.keys=%s", list(session.keys()))
    ruta = request.args.get('ruta', '').strip()

    if not ruta:
        return jsonify({"error": "Parámetro 'ruta' es obligatorio"}), 400

    # Validar que esté dentro de la base permitida
    if not _is_safe_path(ruta, ALLOWED_BASE):
        return jsonify({"error": "Ruta fuera de la ubicación permitida"}), 403

    if not os.path.exists(ruta):
        app.logger.warning("Archivo no encontrado: %s", ruta)
        return jsonify({"error": "El archivo no existe en el servidor"}), 404

    try:
        return send_file(ruta, mimetype='application/pdf', as_attachment=False)
    except Exception:
        app.logger.exception("Error en /api/open_pdf_by_ruta")
        return jsonify({"error": "Error al abrir el PDF"}), 500




# -------------------------FIN-HISTORIAL BOTON--------------- #
# -------------------------------------------------------------------- #
# ---------------------------------MODO BUSQUEDA ----------------------- #
# -------------------------------------------------------------------- #


################FIN DE IMPULSO####################


# -------------------------------------------------------------------- #
# ---------------------------------Excels de vencidos ----------------------- #
# -------------------------------------------------------------------- #


import re
import os
import time
import base64
import mysql.connector
import subprocess
import tempfile
from datetime import datetime, timedelta

# 👉 Añade estos tres imports *antes* de cualquier función que use Alignment, Font, etc.
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from flask import make_response  # asegúrate de tenerlo en tu bloque de imports
import tempfile                 # idem

# --------------------------------------------------------------------
# UTILIDADES NECESARIAS (si no existen aún en tu archivo, copia-pega)
# --------------------------------------------------------------------
# ——— Nueva función para depurar el nombre del abogado ———
def clean_string(text):
    """Elimina caracteres de control no válidos para Excel."""
    if not isinstance(text, str):
        return text
    return re.sub(r'[\000-\010]|[\013-\014]|[\016-\037]', '', text)
def clean_abogado(raw):
    """Depura el campo 'abogado', dejando solo el apellido correcto."""
    if not raw:
        return ""
    s = raw.strip()
    if ';' in s:
        s = s.split(';', 1)[1]
    if ',' in s:
        s = s.split(',', 1)[0]
    return s.strip()

def add_business_days(start_date, business_days):
    """
    Devuelve la fecha resultante tras sumar (o restar) «business_days».
    Si se detecta un rango absurdo (< 1900 o > 2100) devuelve None
    para que el llamador decida ignorar ese registro.
    """
    if business_days == 0:
        return start_date

    step         = 1 if business_days > 0 else -1
    remaining    = abs(business_days)
    current_date = start_date

    for _ in range(remaining):
        current_date += timedelta(days=step)
        while current_date.weekday() >= 5:          # fin de semana
            current_date += timedelta(days=step)

        if current_date.year < 1900 or current_date.year > 2100:
            return None                             # ← descartar

    return current_date



def partition_plazos(plazos):
    acciones, audiencias = [], []
    for f in plazos:
        p = f.get("plazo_atencion")
        if (isinstance(p, int)) or (isinstance(p, str) and p.strip().isdigit()):
            acciones.append(f)
        else:
            audiencias.append(f)
    return acciones, audiencias

def _safe_value(v):
    """Convierte a str y elimina cualquier carácter ilegal para Excel."""
    if isinstance(v, str):
        return clean_string(v)
    return v

def escribir_datos_en_hoja(ws, titulo, headers, filas,
                           start_row: int = 4, start_col: int = 2):
    """
    Escribe título, encabezados y datos en la hoja `ws`, filtrando
    **todas** las cadenas con `_safe_value` para evitar IllegalCharacterError.
    """
    ncols = len(headers)

    # Título centrado
    ws.merge_cells(start_row=3, start_column=start_col,
                   end_row=3, end_column=start_col + ncols - 1)
    c = ws.cell(row=3, column=start_col, value=_safe_value(titulo))
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.font = Font(bold=True)

    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )

    # Encabezados
    for i, h in enumerate(headers):
        ch = ws.cell(row=start_row, column=start_col + i, value=_safe_value(h))
        ch.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ch.font = Font(bold=True)
        ch.fill = PatternFill("solid", fgColor="4F81BD")
        ch.border = thin

    # Datos
    row_xl = start_row + 1
    for fila in filas:
        for j, v in enumerate(fila):
            cell = ws.cell(row=row_xl, column=start_col + j, value=_safe_value(v))
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin
        row_xl += 1

    # Ajuste de anchos
    for i in range(ncols):
        col_letter = get_column_letter(start_col + i)
        max_len = max(
            (len(str(cell.value)) for cell in ws[col_letter] if cell.value), default=0
        )
        ws.column_dimensions[col_letter].width = min(max_len + 2, 30)


def generar_excel_global_modificado(registros, ruta_out):
    acciones, audiencias = partition_plazos(registros)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    if acciones:
        ws = wb.create_sheet("Acciones a tomar")
        escribir_datos_en_hoja(
            ws,
            "RENDICIÓN DE CUENTAS - ABOGADOS PENALES - GENERAL - ACCIONES A TOMAR",
            [
                "Notificación", "ID", "Registro PPU", "Fecha de Recepción por Mesa de Partes",
                "Días de plazo para atención", "Caso fiscal o expediente judicial",
                "Abogado", "Fiscalía", "Situación", "Tipo de acción a tomar", "Fecha Vencimiento"
            ],
            [
                [
                    f"Legajo {f['registro_ppu']} que requiere {f['accion']} de fecha "
                    f"{f['fecha_atencion'].strftime('%d-%m-%Y') if isinstance(f['fecha_atencion'], datetime) else f['fecha_atencion']}",
                    f["id"], f["registro_ppu"], "",
                    f["plazo_atencion"], f["origen"], clean_abogado(f["abogado"]),
                    f["fiscalia"], f["e_situacional"], f["accion"], f["fecha_vencimiento"]
                ] for f in acciones
            ]
        )

    if audiencias:
        ws = wb.create_sheet("Audiencias")
        escribir_datos_en_hoja(
            ws,
            "RENDICIÓN DE CUENTAS - ABOGADOS PENALES - GENERAL - AUDIENCIAS",
            [
                "Notificación", "ID", "Registro PPU", "Fecha de Recepción por Mesa de Partes",
                "Fecha y hora de audiencia", "Caso fiscal o expediente judicial",
                "Abogado", "Fiscalía", "Situación", "Tipo de Audiencia", "Fecha Vencimiento"
            ],
            [
                [
                    f"Legajo {f['registro_ppu']} que tiene audiencia programada para {f['plazo_atencion']}",
                    f["id"], f["registro_ppu"], "",
                    f["plazo_atencion"], f["origen"], clean_abogado(f["abogado"]),
                    f["fiscalia"], f["e_situacional"], f["accion"], f["fecha_vencimiento"]
                ] for f in audiencias
            ]
        )
    wb.save(ruta_out)
from flask import send_file, after_this_request


# --------------------------------------------------------------------
# ENDPOINT → descarga un Excel con plazos NO vencidos
# --------------------------------------------------------------------
@app.route('/api/plazos/no_vencidos_excel', methods=['GET'])
@login_required
def descargar_excel_no_vencidos():
    abogado = request.args.get('abogado', '').strip().upper()
    ahora   = datetime.now()

    # -------- 1. leer BD ----------
    conn = get_db_connection()
    if conn is None:
        return jsonify(error="Error BD"), 500
    try:
        cur = conn.cursor(dictionary=True)
        sql = """
            SELECT dp.*, d.abogado, d.fiscalia, d.origen
            FROM datapenal_plazos dp
            JOIN datapenal d ON dp.registro_ppu = d.registro_ppu
            WHERE (dp.seguimiento IS NULL OR dp.seguimiento <> 'ATENDIDA')
        """
        params = []
        if abogado:
            sql += " AND UPPER(TRIM(SUBSTRING_INDEX(d.abogado,';',-1))) = %s"
            params.append(abogado)
        cur.execute(sql, params)
        filas = cur.fetchall()
    finally:
        cur.close()
        conn.close()

    # -------- 2. filtrar SOLO no-vencidos ----------
    vivos = []
    for f in filas:
        plazo  = f.get('plazo_atencion')
        fecha0 = f.get('fecha_atencion')
        if not (plazo and fecha0):
            continue
        try:
            base = datetime.strptime(str(fecha0)[:10], "%Y-%m-%d") if isinstance(fecha0, str) else fecha0
            if isinstance(plazo, (int, float)) or (isinstance(plazo, str) and plazo.isdigit()):
                due = add_business_days(base, int(plazo))
            elif isinstance(plazo, datetime):
                due = plazo
            elif isinstance(plazo, str) and re.match(r'\d{2}-\d{2}-\d{4}\s+\d{2}:\d{2}', plazo):
                due = datetime.strptime(plazo, "%d-%m-%Y %H:%M")
            else:
                continue
            if (due - ahora).total_seconds() < 0:
                continue  # vencido → fuera
            f['fecha_vencimiento'] = due.strftime("%d-%m-%Y %H:%M")
            vivos.append(f)
        except Exception:
            continue

    if not vivos:
        return jsonify(error="Sin registros no vencidos"), 404

    # -------- 3. generar Excel en memoria ----------
    buf = BytesIO()
    generar_excel_global_modificado(vivos, buf)   # tu helper debe aceptar file-like
    buf.seek(0)                                   # <- muy importante

    # NO cierres buf: déjalo vivo
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="Plazos_NO_Vencidos.xlsx"
    )


@app.route("/api/plazos/vencidos_excel", methods=["GET"])
@login_required
def descargar_excel_vencidos():
    """
    Devuelve un Excel con **todos** los plazos VENCIDOS
    (seguimiento distinto de 'ATENDIDA' y fecha de vencimiento < hoy).

    Puedes filtrar por ?abogado=APELLIDO (se compara contra la parte
    posterior al ';' en d.abogado).
    """
    abogado = request.args.get("abogado", "").strip().upper()
    ahora   = datetime.now()

    # ─ 1. LEER BD ───────────────────────────────────────────────
    conn = get_db_connection()
    if conn is None:
        return jsonify(error="Error BD"), 500

    try:
        cur = conn.cursor(dictionary=True)
        sql = (
            "SELECT dp.*, d.abogado, d.fiscalia, d.origen "
            "FROM datapenal_plazos dp "
            "JOIN datapenal d ON dp.registro_ppu = d.registro_ppu "
            "WHERE (dp.seguimiento IS NULL OR dp.seguimiento <> 'ATENDIDA')"
        )
        params = []
        if abogado:
            sql += " AND UPPER(TRIM(SUBSTRING_INDEX(d.abogado,';',-1))) = %s"
            params.append(abogado)
        cur.execute(sql, params)
        filas = cur.fetchall()
    finally:
        cur.close()
        conn.close()

    # ─ 2. FILTRAR SOLO LOS YA VENCIDOS ──────────────────────────
    vencidos = []
    for f in filas:
        plazo  = f.get("plazo_atencion")
        fecha0 = f.get("fecha_atencion")
        if not plazo or not fecha0:
            continue

        try:
            base = datetime.strptime(str(fecha0)[:10], "%Y-%m-%d") if isinstance(fecha0, str) else fecha0

            if isinstance(plazo, (int, float)) or (isinstance(plazo, str) and plazo.isdigit()):
                due = add_business_days(base, int(plazo))

            elif isinstance(plazo, datetime):
                due = plazo

            elif isinstance(plazo, str) and re.match(r"\d{2}-\d{2}-\d{4}\s+\d{2}:\d{2}", plazo):
                due = datetime.strptime(plazo, "%d-%m-%Y %H:%M")

            else:
                continue

            if due is None:          # ←  ***AÑADE ESTA LÍNEA***
                continue             #     fecha imposible → descarta

        except Exception:
            continue

        if due >= ahora:             # aún no vence ⇒ descarta
            continue

        f["fecha_vencimiento"] = due.strftime("%d-%m-%Y %H:%M")
        vencidos.append(f) 
    if not vencidos:
        return jsonify(error="Sin registros vencidos"), 404

    # ─ 3. GENERAR EXCEL EN MEMORIA ──────────────────────────────
    buf = BytesIO()
    generar_excel_global_modificado(vencidos, buf)   # tu helper acepta file-like
    buf.seek(0)

    # ─ 4. RESPUESTA ─────────────────────────────────────────────
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="Plazos_VENCIDOS.xlsx",
    )

# -------------------------------------------------------------------- #
# ---------------------------------FIN Excels de vencidos ----------------------- #
# -------------------------------------------------------------------- #


# Ejecutar el servidor en producción sin consola (pythonw.exe)
run_server()
